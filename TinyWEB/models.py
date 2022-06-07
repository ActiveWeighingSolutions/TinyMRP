##### Load the init file
#from . import *
## from TinyMRP.models import Part
##print("loaded init")

#Load app and configuration
from TinyWEB import db,app, loadconfiguration, folderout ,fileserver_path, datasheet_folder
from TinyWEB import deliverables_folder, variables_conf, webfileserver
from TinyWEB import maincols, refcols, deliverables, webserver

from flask import (
    Blueprint, flash, g, redirect, session, render_template, request, url_for
)

#Clean this especdially for the upload of bom, if erased traceback will happen
from TinyWEB import *

#To genearte qr codes
import qrcode



#Other libraries
from datetime import datetime, date #for timestamps

import chardet #for getting the encoding of files
import sys, os
import re
from shutil import copyfile
import glob
import pickle #To save bom object session
from pathlib import Path, PureWindowsPath, PurePosixPath
import pandas as pd



import numpy as np
import re
import math

import copy



#PDF libraries
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
import pdfkit #TO EXPORT WEBPAGES TO PDF


# import openpyxl #to manipulate excelfiles
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as openpyxlIm
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from PIL import Image #to process thumbnails


#SQL libraries
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
from sqlalchemy.types import Integer, Float, Boolean,String,Text,NVARCHAR, Date
from flask import render_template, jsonify, request, redirect, url_for, jsonify
from sqlalchemy import create_engine, ForeignKey,select, or_, and_
from sqlalchemy.orm import relationship, backref

#app = Flask(__name__)
#app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
#app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///instance/TinyMRP.db'

#db = SQLAlchemy(app)
  
#db = SQLAlchemy()

from TinyWEB import db, process_conf, variables_conf, property_conf


#To check if conection to file
import urllib

def file_exists(location):

    if "http" in location:
        request = urllib.request.Request(location)
        request.get_method = lambda : 'HEAD'
        try:
            response = urllib.request.urlopen(request)
            return True
        except urllib.error.HTTPError:
            return False
    else:
        if os.path.isfile(location):
            return True
        else:
            return False


def web_to_pdf(url,fileout):
    config= pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
    options = {
                'quiet': '' 
                }
    try:
        pdfkit.from_url(url, fileout, options=options)
    except:
        pass
        #print("Couldn't export to pdf ",url)


#To find the encoding of a particular file
def find_encoding(fname):
    r_file = open(fname, 'rb').read()
    result = chardet.detect(r_file)
    charenc = result['encoding']
    return charenc

def create_folder_ifnotexists(path):    
#Check if outputfolder exists otherwise create it
        foldercheck=os.path.isdir(path)
        if not foldercheck:
            os.makedirs(path)
            return(path)
        else:
            return(path)
        





#To create thumbnails of images
def thumbnail(infile, size=(100, 100)):
    # print("**tumbs infile*****",infile)
    # print(file_exists(infile))

    outfile = os.path.splitext(infile)[0] + ".thumbnail.png"
    if file_exists(outfile):
        # print("**inside if*****")
        if file_exists(infile):
            # print("**inside if file_exists*****")
            if os.path.getatime(infile)>os.path.getatime(outfile):
                try:
                    os.remove(outfile)
                    im = Image.open(infile)
                    im.thumbnail(size, Image.ANTIALIAS)
                    im.save(outfile, "PNG")
                    # print("**tumbs outfile in try*****",outfile)
                    # print(file_exists(outfile))
                    ##print(outfile)
                    return outfile
                except:
                    # print("**tumbs outfile in except*****",outfile)
                    # print(file_exists(outfile))
                    #print("Couldnt update existing OLD thumbnail - ",outfile)
                    return outfile
            else:
                "*** tumbs exists"
                return outfile
        else:
            # print("**tumbs outfile in else*****",outfile)
            # print(file_exists(outfile))
            return(outfile)

    else:
        # print("**inside else*****")
        try:
            im = Image.open(infile)
            im.thumbnail(size, Image.ANTIALIAS)
            im.save(outfile, "PNG")
            
            # print("**tumbs outfile*****",outfile)
            # print(file_exists(outfile))
            return outfile
        except IOError:
            print ("cannot create thumbnail for '%s'" % infile)
            return ""



#To create a QR code to point to part link in tiny:
def qr_code(part):
    # flash(part.partnumber)

    qrfile=fileserver_path+"/Deliverables/png/"+part.partnumber+"_REV_"+part.revision+".qr.jpg"


    if file_exists(qrfile):
        try:
            os.remove(qrfile)
            # flash("erased"+ qrfile)
            
        except:
            # flash("couldnt earse" + qrfile)
            pass

            
    if True:
        image_url="http://"+webserver+"/part/"
        image_url+= part.partnumber+"_rev_"

        if part.revision=="":
                    image_url+= "%25"
        else:
                    image_url+=part.revision

        image_url=image_url.replace(" ","%20")

        #Creating an instance of qrcode

        qr = qrcode.QRCode(
                version=1,
                box_size=10,
                border=5)
        qr.add_data(image_url)
        qr.make(fit=True)

        img = qr.make_image(fill='black', back_color='white')
        tempfile=qrfile.replace('.qr.jpg','.temp.jpg')
        tempfile=img.save(qrfile)
        img.close()

        return qrfile

    if False:
        flash("Problem creating QR for " + part.partnumber)
        return None






    


#To find the children based on a flatbom and a bom per the class definition below of solidbom
def get_children(father_partnumber,father_rev,bom,flatbom, qty="total"):
        children=bom.loc[(bom['father_partnumber']==father_partnumber) & (bom['father_revision']==father_rev)]
        children_rename_dict={}; children_rename_dict['child_partnumber']='partnumber'
        children_rename_dict['child_revision']='revision'
        children=children.rename(columns=children_rename_dict)

        
        children_flatbom=flatbom.merge(children,on=['partnumber','revision'],
                          how='left',indicator=True).query('_merge == "both"').drop(columns='_merge').reset_index(drop=True).sort_values(by='partnumber')

        if len(children)>0:
            return children_flatbom
        else:
            return []
    
        




class User(db.Model):
    # Defines the Table comment
    __tablename__ = "user"

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    username = db.Column(db.String,unique=True , nullable=False)
    password = db.Column(db.String, nullable=False)
    email = db.Column(db.String, nullable=False)
    password = db.Column(db.String)
    authenticated = db.Column(db.Boolean, default=False)

    def is_active(self):
        """True, as all users are active."""
        return True

    def get_id(self):
        """Return the email address to satisfy Flask-Login's requirements."""
        return self.email

    def is_authenticated(self):
        """Return True if the user is authenticated."""
        return self.authenticated

    def is_anonymous(self):
        """False, as anonymous users aren't supported."""
        return False


class Comment(db.Model):
    # Defines the Table comment
    __tablename__ = "comment"
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    part_id = db.Column(db.Integer, ForeignKey('part.id'), nullable=False)
    user_id = db.Column(db.Integer,ForeignKey('user.id'), nullable=False)
    body =db.Column(db.String)
    category =db.Column(db.String)
    created=db.Column(db.Date)
    pic_path=db.Column(db.String)



    
    def __init__(self, part_id="",user_id="",body="",category="", created="",pic_path=""):
        self.part_id = part_id
        self.user_id = user_id
        self.body=body
        self.category =category
        self.pic_path=pic_path
        self.created=created

class solidbom():
    
    def __init__(self, filein,deliverableslocation,outputfolder,toppart=None):
            
            ###Load configuration
            [self.process_conf,self.property_conf, self.variables_conf] = loadconfiguration()
        
            ### List of invalid chars, names, etc...
            ### this is to be set up from a conf file or something
            self.file=filein
            self.invalid_col_chars=['.','-',' ']
            self.renamelist={'part number':'partnumber',
                        'sw-configuration name(configuration name)':'sw_configuration',
                        'sw-folder name(folder name)':'folder',
                        'sw-file name(file name)':'file',
                        'item no.':'item_no'}
            #to rename teh cols
            self.renamedict={}
            
            #to drop the cols that are not required
            self.dropcols=[]
            
            self.col_clean_list=[]
            
            #Variable types
            self.int_cols=['qty']
            self.float_cols=['mass','thickness']
            self.bool_cols=['spare_part']
            
            #Add timestamp of bom creation
            self.timestamp=date.today().strftime('%d_%m_%Y')
            
                       
            #Default folder out
            self.folderout=outputfolder
            self.deliverables_folder=deliverableslocation
            
            
            ##### Load the input file and create the dataframe
            if filein!="":
                file_enc=find_encoding(filein)
                self.filedata=pd.read_csv(filein, encoding =file_enc,sep='\t', lineterminator='\r',dtype=str)
                self.data=self.filedata.copy()
                
                ################################################
                ######################################################
                #Apply company customization 
                self.customize()
                
                #Clean the entry data
                self.clean_data()
                
                #Get the root component definition
                self.root_definition()
                
                
                            
                #Create bom objects
                self.createbom()
                
                #Screening based on custom properties
                self.property_screening()
                
                #Find related files/deliverables
                self.find_deliverables()
            
                #Dataframe to database
                self.createdatabase()
                
            elif toppart!=None:
                
                self.part=toppart
                self.partnumber=toppart.partnumber
                self.revision=toppart.revision
                self.description=toppart.description
                #self.flatbom=get_flatbom(self.partnumber,self.revision)
                self.tag=self.partnumber+"_REV_"+self.revision#+"-"#+self.timestamp


    def solidbom_from_flatbom(object_list,part,outputfolder="",sort=False):
        if outputfolder=="":
            folderout="temp"
        else:
            folderout=outputfolder
        #print("---obj in--- for " , part.partnumber)
        # for pepe in object_list:
        #     #print(pepe.partnumber)
        #     pass
        bomout=solidbom ( "",deliverables_folder,folderout+"/" , toppart=part)

        bomout.flatbom=pd.DataFrame([x.as_dict(folder=fileserver_path) for x in object_list])

        # #print("---bom out --")
        # for i,row in bomout.flatbom.iterrows():
        #     #print(row.partnumber)
        return bomout
        #print("---end --")                
        #print(" ")                
            
    def find_deliverables(self):

        #Create the empty cols for storing the info
        for extension in deliverables:
            self.flatbom[extension]=False
            self.flatbom[extension+'path']=''
        
        
        for i in range(len(self.flatbom)):
            partstring=self.flatbom.at[i,'file']+"_REV_"+self.flatbom.at[i,'revision']
            
            for extension in deliverables:
                extension_folder=self.deliverables_folder+extension+"/"
                file_string=extension_folder+partstring+"."+extension
                if os.path.isfile(file_string): 
                    self.flatbom.at[i,extension+'path']=file_string
                    self.flatbom.at[i,extension]=True
            
       
             
    def root_definition(self):
        self.partnumber=self.data.at[0,'partnumber']
        self.revision=self.data.at[0,'revision']
        self.description=self.data.at[0,'description']
        
        self.tag=self.partnumber+"_REV_"+self.revision#+"-"+self.timestamp
        
        #Set output location
        self.folderout=self.folderout+self.tag+"/"
        
        create_folder_ifnotexists(self.folderout)
        
        
             
    def customize(self):
        
           
        #Build the rename dictionary and rename cols
        for prop in property_conf.keys():
            self.renamedict[property_conf[prop]['custom_property']]=prop
        
        self.data=self.data.rename(columns=self.renamedict) 
        
        #Drop the non required properties from the dataframe
        for col in self.data.columns:
            if not col in [*property_conf]: 
                self.dropcols.append(col)
        self.data= self.data.drop(self.dropcols, axis = 1)
        
       
        
      
             
           
    def clean_data(self):
        
        ###Remove all the entries with no file related to it
        ###This will hasve to be revised if the cut list and sheet metal properties
        ### are to be counted
        self.data.dropna(subset = ["file"], inplace=True)


        #Transform to right data type based on init lists:
        #Nan values will be filled with empty string or 0
        for col in self.data:
            
            if col in self.int_cols:
                try:
                    self.data[col]=pd.to_numeric(self.data[col].fillna(0).astype(int))
                except:
                    #print("Problems tranforming to int " , col)
                    #print("forcing and replacing non valid with 0.0, check impact on output files")
                    self.data[col]=pd.to_numeric(self.data[col], errors='coerce').fillna(0).astype(int)
                
            elif col in self.float_cols:
                try:
                    self.data[col]=pd.to_numeric(self.data[col].fillna(0.0).astype(float))
                except:
                    #print("Problems tranforming to float " , col)
                    #print("forcing and replacing non valid with 0.0, check impact on output files")
                    self.data[col]=pd.to_numeric(self.data[col], errors='coerce').fillna(0.0).astype(float)
            elif col in self.bool_cols:
                self.data[col]=self.data[col].fillna(False).astype(bool)
            else:
                self.data[col]=self.data[col].fillna("").astype(str)
            
        
        #Remove empty rows, spaces and and dirty data
        self.data['revision']=self.data['revision'].astype(str)
        #Remove non alfanumeric from revision and replace by "" (in case empty or funny)
        self.data.loc[self.data.revision.str.isalnum()==False,'revision']=''
        self.data.loc[self.data.revision=="nan",'revision']=''
        
        #Remove leading spaces in part number
        self.data['partnumber']=self.data['partnumber'].str.lstrip()
        
        
        #Reorder the columns for easier debugging and manipulations
        newcols=maincols+refcols
        for col in self.data.columns.to_list():
            if not col in newcols: newcols.append(col)
        
        self.data=self.data.reindex(columns=newcols)
        
        
        #Put process, finish and treatment as lowercase
        for prop in lowercase_properties:
            self.data[prop]=self.data[prop].str.lower()
      
        #Remove all non desired properties
        

         
        

    def property_screening(self):

              
        #Account for the hardware in process related
        for folder in hardware_folder:
            self.flatbom.loc[self.flatbom['folder'].str.lower().str.contains(folder),['process','process2','process3']]=['hardware',"",""]
       
        
       
  
        
    def createdatabase(self):
        
        
        #Add parts to database
        #The overwrite must be revised
        for index, row in self.flatbom.iterrows():
            database_part=Part()
            database_part.partfromlist(row)

            #Find if the part is already in database
            check_part=db.session.query(Part).filter(and_(Part.partnumber==row['partnumber'] ,
                                                     Part.revision==row['revision'])
                                                     ).first()
            #Create or update qrcode
            qr_code(database_part)

            #IF part exists overwrite attributes
            if check_part==None:
                db.session.add(database_part)
                

            else:
                for prop in property_conf.keys():
                    if hasattr(database_part, prop):
                        setattr(check_part,prop,getattr(database_part,prop))
                        
      
        #Commit changes on part table
        db.session.commit() 
        db.session.close()
        


        #Add bom to database
        for index, row in self.flatbom.iterrows():            
        
            database_part=Part()
            database_part=db.session.query(Part).filter(and_(Part.partnumber==row['partnumber'],
                                                     Part.revision==row['revision'])
                                                     ).first()

                        
            #Erase all bom entries of related part downstrema, erase kids entries
            if database_part.partnumber!="":
                bomentries=db.session.query(Bom).filter(Bom.father_id==database_part.id)
                for bomline in bomentries:
                    db.session.delete(bomline)
                db.session.commit() 

                #Get children from solidbom
                children=get_children(database_part.partnumber,database_part.revision,self.bom, self.flatbom)

                #Add children to database
                if len(children)>0:
                    for i, childrow in children.iterrows():
                        childpart=db.session.query(Part).filter(and_(Part.partnumber==childrow['partnumber'] ,
                                                     Part.revision==childrow['revision'])
                                                     ).first()
                        if childpart!=None:
                            bomentry=Bom(database_part.id,childpart.id,childrow['qty'])
                            db.session.add(bomentry)

        #Commit changes on bom table
        db.session.commit() 
        db.session.close()


        
    def createbom(self):

        self.flatbom=self.data.copy()

        #Remove qty and tree reference to  get unique entries
        #And add the totalquanty
        del self.flatbom['item_no']
        del self.flatbom['qty']
        self.flatbom['totalqty']=0
            
        # #Drop flatbom duplicates
        self.flatbom= self.flatbom.drop_duplicates()
        self.flatbom=self.flatbom.reset_index(drop=True)
        
        
        ## Create bom dataframe
        self.bom= pd.DataFrame({'father_partnumber':"",   
                'father_revision':"",
                'child_partnumber':self.data['partnumber'],
                'child_revision':self.data['revision'],
                'qty':self.data['qty'],
                'ref': self.data ['item_no']
                })
        
        #self.bom.reset_index(drop=True)

        #Filter bom ref for pointing only to father for 
        #accouting of duplicated configurations
        for index, row in self.data.iterrows():
            self.bom.at[index,'ref']=re.sub(r"\..?[0-9]$", "",self.bom.at[index,'ref'])
        
        
        #Build bom table finding referenced father part number
        for index, row in self.data.iterrows():
                
                temp=self.data.loc[self.data['item_no'] ==re.sub(r"\..?[0-9]$", "", row ['item_no'])].reset_index(drop=True)
                self.bom.at[index,'father_partnumber']=temp.at[0,'partnumber']
                self.bom.at[index,'father_revision']=temp.at[0,'revision']
                #If couldnt repalce, it means twe are in the top level and we add the mark root
                if self.bom.at[index,'father_partnumber']==self.bom.at[index,'child_partnumber']:
                    self.bom.at[index,'father_partnumber']='root'
  
       #Combine duplicate entries quantities and reset index
        self.bom=self.bom.groupby(['child_partnumber','child_revision','father_partnumber','father_revision','ref'])['qty'].sum().to_frame().reset_index()
        self.bom.reset_index(drop=True)
        
        
      
         # #Compute the total quantity for each partnumber
        self.data['branchqty']=self.data['qty']
        self.data=self.data.set_index('item_no')

    
        for index, row in self.data.iterrows():
            father_index=re.sub(r"\..?[0-9]$", "", index)
            while len(father_index.split('.'))>1:
                self.data.at[index,'branchqty']=self.data.at[index,'branchqty']*self.data.at[father_index,'branchqty']
                father_index=re.sub(r"\..?[0-9]$", "", father_index)
            
        for index, row in self.flatbom.iterrows():
            
            self.flatbom.at[index,'totalqty']=self.data[(self.data['partnumber']==row['partnumber'] )& (self.data['revision']==row['revision'])]['branchqty'].sum()
        
 
        
    def solidbom_to_excel(self,process="",consumed=[],hardware=[]):
        redFill = PatternFill(start_color='FFEE1111',
                      end_color='FFEE1111',
                      fill_type='solid')
        yellowFill = PatternFill(start_color='00FFFF00',
                              end_color='00FFFF00',
                              fill_type='solid')
        
        bom_in=self.flatbom.loc[(self.flatbom['process']!='hardware') & (self.flatbom['process2']!='hardware') & (self.flatbom['process3']!='hardware')]
        if hardware==[]:
            hardware_flatbom=self.flatbom.loc[(self.flatbom['process']=='hardware') | (self.flatbom['process2']=='hardware') | (self.flatbom['process3']=='hardware')]
        consumed_flatbom=consumed

        #Standard columns and sheet defionitons
        std_columns=['Screenshot','partnumber','qty','approved','material','process','finish','category','supplier','colour']
        hardware_columns=['Screenshot','partnumber','qty','approved','material','finish','supplier','supplier_partnumber']


        #Prepare file and format wrap
        if process!="":
            self.excel_file=self.folderout +process.upper()+"-Table-" +self.tag  +".xlsx"

            
            # std_columns=['Screenshot','partnumber','qty','material']
        else:
            self.excel_file=self.folderout + "BOM_tables-" +self.tag  +".xlsx"

        writer = pd.ExcelWriter(self.excel_file, engine = 'xlsxwriter')
        workbook = writer.book
        wrap_format = workbook.add_format({'text_wrap': True})

        
        sheetpairs=[['Flatbom',bom_in,std_columns],['Consumed',consumed_flatbom,std_columns],['Hardware',hardware_flatbom,hardware_columns]]
        sheetpairs=[[x,y,z] for x,y,z in sheetpairs if len(y)>0]
        # print(sheetpairs)

        print("SHEETPAIRS",len(sheetpairs))
        
        for sheet_name,bomflat,columns in sheetpairs: 
            print(sheet_name,len(bomflat),list(bomflat),columns) 
            # print("BOMFLAT",len(bomflat),type(bomflat))  
            # print(bomflat)
            # if process=="cutfold":
            #     bomflat=self.flatbom.loc[(self.flatbom['process']=='lasercut')] # | (self.flatbom['process2']=='lasercut') | (self.flatbom['process3']=='lasercut') |
                #                          (self.flatbom['process']=='folding') | (self.flatbom['process2']=='folding') | (self.flatbom['process3']=='folding') ]


            #Rearrange the cols
            notmain_cols=[]
            last_cols=[]
            
            for col in bomflat:
                if col in maincols or col in refcols:
                    pass
                elif "path" in col or col in deliverables or col in ['eprt','edrw','easm','threemf']:
                    last_cols.append(col)
                elif col!="_sa_instance_state":
                    notmain_cols.append(col)

            
            #Add screenshot col
            bomflat["Screenshot"]=""
            
            #Copy of bom for bom generation only (avoid having images path collumns)
            bom_image=bomflat[['partnumber']+['revision']+['pngpath']+['png']]

            #Replace datasheet location by web link:
            bomflat['datasheet']=bomflat['datasheet'].str.replace("//","/")
            bomflat['datasheet']=bomflat['datasheet'].str.replace(fileserver_path,"http://"+webfileserver)


            #Combine all the processes in one col
            # bomflat['process']=bomflat['process']+ " - "+bomflat['process2']+ " - "+bomflat['process3']

            #Replace partnumber by supplierpartnumber in for purchased items:
            # and combine the partnubmer rev and description 
            for index, row in bomflat.iterrows():
                
                if 'oem' in row['partnumber'].lower():
                    
                    bomflat.at[index,'partnumber']=row['supplier_partnumber'] + " - "+row['description']
                else:
                    bomflat.at[index,'partnumber']=row['partnumber']+" REV "+row['revision']+ " - "+row['description']

                bomflat.at[index,'process']=row['process']+" - "+row['process2']+ " - "+row['process3']
                    


            #Combine oem and supplier
            
            bomflat['supplier']=bomflat['oem']+" "+bomflat['supplier']

            


            if process!="" and process!='cutlist' and process!="cutfold"and process!="steel":
                bomflat=bomflat[["Screenshot"]+process_conf[process]['fields']]
                
            else:
                bomflat=bomflat[["Screenshot"]+maincols+refcols+notmain_cols+last_cols]
        



        
            #wrap_format.set_border(6)
            
            # #Dump dataframe to excel
            if process!="":
                sheet_name = process.upper()+ ' scope of supply'
        
            
            #Add col custmization for cutlist:
            if process=='cutlist':
                columns=['Screenshot','partnumber','qty','approved','material','thickness','finish','process']



            #Add missing cols TO AVOID ERRORS
            
            finalcols=[]
            for col in columns: 
                if col in list(bomflat):
                    finalcols.append(col)
                    print("YES",col)
                else:
                    print("NO",col)
            

        
            #Reset index and Increment so it starts with 1
            bomflat=bomflat.reset_index(drop=True)
            bomflat.index += 1


            # bomflat.to_excel(writer, sheet_name = sheet_name, columns=['partnumber','revision','qty'])
            bomflat.to_excel(writer, sheet_name = sheet_name, columns=finalcols)
            
            

            # Set the worksheet and  autofilter.
            worksheet = writer.sheets[sheet_name]
            (max_row, max_col) = bomflat[finalcols].shape
            worksheet.autofilter(0, 0, max_row, max_col )
            worksheet.set_column(2,max_col,None,wrap_format)

            

            
            #Width of cols
            worksheet.set_column(0,0,4,wrap_format)
            worksheet.set_column(1,1,8,wrap_format)
            worksheet.set_column(2,2,60,wrap_format)
            worksheet.set_column(3,4,4,wrap_format)
            
            worksheet.set_column(5,max_col,20,wrap_format)




            # Add images section
            
            i=-1           
            for index, row in bom_image.iterrows():
                i=i+1
                thumb=thumbnail(row['pngpath'])
                #print(thumb) 

                #Adjust row height
                worksheet.set_row(i+1,30)

                #Add image
                cell='B'+str(i+2)

                image_url="http://"+webserver+"/part/"
                image_url+= row['partnumber']+"_rev_"

                if row['revision']=="":
                    image_url+= "%25"
                else:
                    image_url+= row['revision']

                #print(row['png'])

                    

                if (row['png']!="FALSE" and  row['png']!=False ) or thumb!="":
                    try:
                        worksheet.insert_image(cell, thumb, {'x_offset': 1,
                                                        'y_offset': 1,
                                                        'x_scale': 0.5,
                                                        'y_scale': 0.5,
                                                        'object_position': 1,
                                                        'url': image_url})
                    except:
                        print("***** ISSUES WITH EXCEL IMAGE FOR ***", row['partnumber'])
                

                
        #Close workbook with xlsxwriter so openpyxl can open it
        workbook.close()

        return self.excel_file
        
        

        
    def get_parents(self,partnumber,revision):
        parents=self.bom.loc[(self.bom['child_partnumber']==partnumber) & (self.bom['child_revision']==revision)]
        parents_rename_dict={} ; parents_rename_dict['father_partnumber']='partnumber'
        parents_rename_dict['father_revision']='revision'
        parents=parents.rename(columns=parents_rename_dict)
        
        parents=self.flatbom.merge(parents,on=['partnumber','revision'],
                          how='left',indicator=True).query('_merge == "both"').drop(columns='_merge').reset_index(drop=True)
        
        if len(parents)>0:
            return parents
        else:
            return []
        
        
    def gather_datasheet(self):
        
        #Check if outputfolder exists otherwise create it
        outputfolder=self.folderout+"datasheets/"
        create_folder_ifnotexists(outputfolder)
        
        if len(self.flatbom[(self.flatbom['process']=='purchase') |( self.flatbom['process2']=='purchase')|( self.flatbom['process3']=='purchase')]):
            purchasefolder=self.folderout+"purchase/"
            create_folder_ifnotexists(purchasefolder)
            
        for i in range(len(self.flatbom)):
            
            targetfile=outputfolder + self.flatbom['partnumber'][i]+"-datasheet"
    
            if os.path.isfile(self.flatbom.at[i,"datasheet"]):
                sourcefile=self.flatbom["datasheet"][i]
                fileName, fileExtension = os.path.splitext(sourcefile)
                

                
                if "png" in fileExtension.lower() or "jpg" in fileExtension.lower() : 
                    try:
                        image1=Image.open(sourcefile)
                        im1=image1.convert('RGB')
                        targetfile=targetfile+".pdf"
                        im1.save(targetfile)
                        self.flatbom.at[i,'datasheet']=targetfile
                    except:
                        print (self.flatbom['partnumber'][i] + " PROBLEMS COMPILING DATASHEET ", sourcefile)
                elif "pdf" in fileExtension.lower():    
                    targetfile=targetfile + fileExtension
                    sourcefile=PureWindowsPath(sourcefile)
                    targetfile=PureWindowsPath(targetfile)
                    ##print(sourcefile, " copy to ", targetfile)
                    try:
                        copyfile(sourcefile,targetfile)
                        
                        self.flatbom.at[i,'datasheet']=targetfile
                    except:
                        print (self.flatbom['partnumber'][i] + " PROBLEMS COMPILING DATASHEET ", sourcefile, targetfile)
                        
                if self.flatbom['process'][i]=='purchase' or self.flatbom['process2'][i]=='purchase' or self.flatbom['process3'][i]=='purchase' :
                        purchasefile=purchasefolder + self.flatbom['partnumber'][i]+".pdf"
                        try:
                            copyfile(targetfile,purchasefile)
                        except:
                            #print("Couldn't copy ", purchasefile  , " to ", targetfile)
                            pass
                    
            elif self.flatbom.at[i,"datasheet"]!="":
                    try:
                        web_to_pdf(self.flatbom['link'][i],targetfile+".pdf")
                        #print(self.flatbom['partnumber'][i] , " DATASHEET FROM WEB")
                        
                    except:
                        pass
                        # try:
                        #     #print(self.flatbom['partnumber'][i] , " NO DATASHEET - ", sourcefile)
                        # except:
                        #     #print(self.flatbom['partnumber'][i] , " invalid SOURCEFILE !!!!!!!!!!")
                        # # self.flatbom.at[i,'notes'].append("Invalid datasheet source")
    
    def gather_deliverables(self):
        
        #Loops over the processes in the excel configurator file 
        #but excludes the last one (others) that is why [*process_conf][:-1] 
        # so be careful when adding more
        
        for process in [*process_conf][:-1]:
            bom_in=self.flatbom.loc[(self.flatbom['process']==process) | (self.flatbom['process2']==process) | (self.flatbom['process3']==process)]
            
            if len(bom_in)==0:
                continue
            
            musthave=[]
            if process_conf[process]['pdf']==1: musthave.append("pdf")
            if process_conf[process]['dxf']==1: musthave.append("dxf")
            if process_conf[process]['step']==1: musthave.append("step")
            
            #Check if outputfolder exists otherwise create it
            outputfolder=self.folderout+process+"/"
            
            #print(outputfolder)
            if len(musthave)>0:create_folder_ifnotexists(outputfolder)
    
            
            for index, row in bom_in.iterrows():
                ##print(index,row['partnumber'], row["pdfpath"])  
                
                filenamebit=row["partnumber"]+"_REV_"+row["revision"]
                for extension in musthave:
                    sourcefile=self.deliverables_folder+extension+"/"+filenamebit+"."+extension
                    
                    if process=='folding' or process=='lasercut' or process=='profile cut':
                        targetfile=outputfolder+filenamebit
                        targetfile=targetfile+"-"+row["material"]+"_"+str(row["thickness"])+"mm."+extension
                        # if row["thickness"]==0 or str(row["thickness"])=='':
                           # self.flatbom.at[index,'notes'].append("Missing thickness")
                        
                    else:
                        targetfile=outputfolder+filenamebit+"."+extension
                    
                                        
                    if os.path.isfile(sourcefile):
                        copyfile(sourcefile,targetfile)
                #print(sourcefile,targetfile)

  
class Bom(db.Model):
    # Defines the Table Name user
    __tablename__ = "bom"
    
	# Defines the variables
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    father_id = db.Column(Integer, ForeignKey('part.id'), nullable=False)
    child_id = db.Column(Integer, ForeignKey('part.id') , nullable=False)
    qty=db.Column(Integer, nullable=False)
        
    
    def __init__(self, father_id, child_id,qty):
        self.father_id = father_id
        self.child_id = child_id
        self.qty = qty
        self.child=Part.query.filter_by(id=self.child_id).first()
        
    def getchild(self):
        self.child=Part.query.filter_by(id=self.child_id).first()
        return self.child
        
    def __repr__(self):
        
        self.getchild()
        return f'BOM( {self.child.partnumber}, {self.child.revision} , quantity {self.qty})'
    def __str__ (self):
        self.getchild()
        return f'BOM( {self.child.partnumber} , {self.child.revision} ,quantity {self.qty})'    
             
class Part(db.Model):
    # Defines the Table Name user
    __tablename__ = "part"
    
	# Makes three columns into the table id, name, email
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    partnumber = db.Column(db.String, nullable=False)
    revision = db.Column(db.String)
    
    approved = db.Column(db.String)
    author = db.Column(db.String)
    category = db.Column(db.String)
    configuration = db.Column(db.String)
    colour = db.Column(db.String)
    datasheet = db.Column(db.String)
    description = db.Column(db.String)
    drawndate = db.Column(db.String)
    file = db.Column(db.String)
    finish = db.Column(db.String)
    folder = db.Column(db.String)
    link = db.Column(db.String)
    mass = db.Column(db.Float)
    material = db.Column(db.String)
    oem = db.Column(db.String)
    
    process = db.Column(db.String)
    process2 = db.Column(db.String)
    process3 = db.Column(db.String)
    
    spare_part = db.Column(db.Boolean)
    supplier = db.Column(db.String)
    supplier_partnumber = db.Column(db.String)
    thickness = db.Column(db.Float)
    treatment = db.Column(db.String)
    colour = db.Column(db.String)
    notes = db.Column(db.String)
    asset = db.Column(db.String)
    
    children= relationship("Part", 
                           secondary="bom",
                           primaryjoin=id==Bom.father_id,
                           secondaryjoin=id==Bom.child_id,
                           backref="parent" )
    comments=relationship("Comment")

    totalqty = db.Column(db.Integer)
    qty = db.Column(db.Integer)

    def as_dict(self,folder=""):
        # if not hasattr(self,'qty'):
        #     self.qty=0
        #     self.totalqty=self.qty
        # else:
        #     if not hasattr(self,'totalqty'):
        #         self.totalqty=0

        self.datasheet_available=False
        self.png=False
        self.pdf=False
        self.eprt=False
        self.edrw=False
        self.easm=False
        self.edr=False
        self.png_d=False
        self.dxf=False
        self.step=False
        self.threemf=False
        
        self.datasheet_link=""   
        self.modelpath=""
        self.pngpath=""
        self.pdfpath=""
        self.dxfpath=""
        self.edrpath=""
        self.steppath=""
        self.threemfpath=""
        self.png_dpath=""

        if folder!="":
            self.updatefilespath(folder,local=True)

        # tempdict=copy.deepcopy(self.__dict__)
        # tempdict=tempdict.pop('_sa_instance_state',None)
        dicto=self.__dict__
        outdict={}

        for key in dicto.keys():
            if key!='_sa_instance_state':
                outdict[key]=dicto[key]
        return outdict        
        
        # # return tempdict
        # return self.__dict__

    def partFromDict(dicto):
            if type(dicto)==Part:
                return dicto
            else:
                parto=Part.query.filter_by(partnumber=dicto['partnumber'],revision=dicto['revision']).first()
                parto.qty=dicto['qty']
                return parto


        
    
    def getchildren (self):
        outlist=[]
        
        for child in self.children:
            kid={}
            kid['part']=child
            kid['qty']=Bom.query.filter_by(father_id=self.id,child_id=child.id).first().qty
            outlist.append(kid)
        return outlist
    
    def children_with_qty (self):
        outlist=[]
        
        for child in self.children:
            child.qty=Bom.query.filter_by(father_id=self.id,child_id=child.id).first().qty
            outlist.append(child)

        return outlist

    def get_tag(self,addtimestamp=True):
        
        self.tag=self.partnumber+"_REV_"+self.revision
        if addtimestamp:
            self.tag=self.tag+"-"+date.today().strftime('%d_%m_%Y')
        return self.tag


    def parents_with_qty (self):
        outlist=[]
        
        for parent in self.parent:
            parent.qty=Bom.query.filter_by(father_id=parent.id,child_id=self.id).first().qty
            outlist.append(parent)

        return outlist

    def hasProcess(self,process):
        if process in self.process or process in self.process2 or process in self.process3:
            return True
        else:
            return False
    
             



    def hasConsumingProcess(self):
        processlist=[self.process,self.process2,self.process3]
        
        processlist=[x for x in processlist if x in process_conf.keys()]

        consume=False

        for process in processlist:
            if int(process_conf[process]['priority'])<20:
                consume=True

        return consume



    def isMainProcess(self,process):
        mainprocess_bool=False

        if self.hasProcess(process):
            processlist=[self.process,self.process2,self.process3]
            processlist=[x for x in processlist if x in process_conf.keys()]
            if len(processlist)>0:
                for x in processlist:
                    if int(process_conf[process]['priority'])<= int(process_conf[x]['priority']):
                        mainprocess_bool=True
                        
                    else:
                        mainprocess_bool=False
            

        return mainprocess_bool


    def MainProcess(self):
        processlist=[self.process,self.process2,self.process3]
        processlist=[x for x in processlist if x in process_conf.keys()]

        mainprocess=self.process
        if len(processlist)>0:
            for x in processlist:

                if int(process_conf[x]['priority'])<= int(process_conf[mainprocess]['priority']):
                    mainprocess=x

        else:
            mainprocess='other'
        self.mainprocess=mainprocess
                
        return mainprocess



    def __repr__(self):
        return f'part( {self.partnumber} , {self.revision} , {self.description})'
    def __str__ (self):
        return f'part( {self.partnumber} , {self.revision} , {self.description})'

    def updatefilespath(self,folderin,local=False, png_thumbnail=False):
        #This functions check first in the file server local path and then adds the http link
        folder=folderin
        # folder=webfileserver

        self.tag=self.partnumber+"_REV_"+self.revision

        pngfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"png/"
        pdffolder=fileserver_path+variables_conf['deliverables_folder']['value']+"pdf/"
        edrfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"edr/"
        stepfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"step/"
        dxffolder=fileserver_path+variables_conf['deliverables_folder']['value']+"dxf/"
        threemffolder=fileserver_path+variables_conf['deliverables_folder']['value']+"3mf/"

        #Ssame for datasheets
        datasheetfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"/datasheet/"
        
        self.pdfpath=pdffolder+self.tag+".pdf"
        self.steppath=stepfolder+self.tag+".step"
        self.dxfpath=dxffolder+self.tag+".dxf"
        self.eprtpath=edrfolder+self.tag+".eprt"
        self.easmpath=edrfolder+self.tag+".easm"
        self.edrwpath=edrfolder+self.tag+".edrw"
        self.png_dpath=pngfolder+self.tag+"_DWG.png"
        self.pngpath=pngfolder+self.tag+".png"
        self.qrpath=pngfolder+self.tag+".qr.jpg"
        self.threemfpath=threemffolder+self.tag+".3mf"

        if self.hasProcess("hardware") or self.hasProcess("purchase") and not file_exists(self.pngpath):
            self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"
  
        #if self.hasProcess("hardware") and not file_exists(self.pngpath):
        #    if not local:
        #        self.pngpath=pngfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".png"
        #    else:
        #        self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"

        #Adding boolean value to false for all files
        self.datasheet_available=False
        self.png=False
        self.qr=False
        self.pdf=False
        self.eprt=False
        self.edrw=False
        self.easm=False
        self.edr=False
        self.png_d=False
        self.dxf=False
        self.step=False
        self.threemf=False
        
        
        #Add directly the link for datasheets withouth checking file exists
        
        if self.datasheet:
            ##print(self.datasheet)
            
            self.datasheet.replace('"', '')
            self.datasheet.replace("file:///",'')
            self.datasheet.replace('%20',' ')
            # path, filename = os.path.split(self.datasheet)
            # flash(PureWindowsPath(self.datasheet).name)
            # flash(filename)
            filename=PureWindowsPath(self.datasheet).name
            self.datasheet=datasheetfolder+filename
            
            if file_exists(self.datasheet):
                self.datasheet_link=folder+variables_conf['deliverables_folder']['value']+"/datasheet/"+ filename
                self.datasheet_link=self.datasheet_link.replace(' ','%20')
                self.datasheet_available=True
            else:
                if file_exists(self.datasheet+".pdf"):
                     self.datasheet_link=folder+variables_conf['deliverables_folder']['value']+"/datasheet/"+ filename+".pdf"
                     self.datasheet_link=self.datasheet_link.replace(' ','%20')
                     self.datasheet_available=True
            #if hasattr(self,"datasheet_link"): 
            #    #print(self.datasheet_link)
      
           



        #UPdate the folder location to the given in function, normally if not local
        #the folder path will be the fileserver address
        if not local:
            pngfolder=folder+variables_conf['deliverables_folder']['value']+"png/"
            pdffolder=folder+variables_conf['deliverables_folder']['value']+"pdf/"
            edrfolder=folder+variables_conf['deliverables_folder']['value']+"edr/"
            stepfolder=folder+variables_conf['deliverables_folder']['value']+"step/"
            dxffolder=folder+variables_conf['deliverables_folder']['value']+"dxf/"
            threemffolder=folder+variables_conf['deliverables_folder']['value']+"3mf/"

        #To reduce the image quality on the flatbom if needed (too much bandwidth)
        
        
        #Png path
        

        ##print(self.pngpath)
        if file_exists(self.pngpath):
            # #print("exist")
            if png_thumbnail:
                pngfolder=fileserver_path+variables_conf['deliverables_folder']['value']+"png/"
                self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"
                self.pngpath=thumbnail(self.pngpath)
                
                try:
                    path, filename = os.path.split(self.pngpath)
                    
                    self.pngpath=folder+variables_conf['deliverables_folder']['value']+"png/"+filename
                    
                    self.png=True 
                    
                except:
                    #print("Problems with ", self.partnumber)
                    #print("pngfolder  ", fileserver_path)
                    #print("folder  ", folder)
                    #print("fileserver_path  ", fileserver_path)
                    
                    self.png=False 
            else:
                self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"
             

            if not local:
                self.pngpath=self.pngpath.replace(' ','%20')
  
        else:
            try:
                self.pngpath=pngfolder+self.file+"_REV_"+self.revision+".png"
                self.png=True 
            except:
                pass
    

        
        
        
        if self.folder!= None and self.file!= None:

            #Model paths added directly
            self.modelpath=self.folder +self.file+ ".SLDPRT"
            if file_exists( self.modelpath.upper()+ ".SLDPRT"): 
                self.modelpath= self.modelpath+ ".SLDPRT" 
            if file_exists( self.modelpath+ ".SLDASM"): 
                self.modelpath= self.modelpath+ ".SLDASM" 

                    
            # #qr path# self.qrpath=qr_code(self)
            self.qrpath=pngfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".qr.jpg"
            self.qr=True 





            #png_d path
            if file_exists(self.png_dpath):
                self.png_dpath=pngfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+"_DWG.png"
                self.png_d=True 
        

            #pdf path
            if file_exists(self.pdfpath):
                self.pdfpath=pdffolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".pdf"
                self.pdf=True 


            #dxf path
            if file_exists(self.dxfpath):
                self.dxfpath=dxffolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".dxf"
                self.dxf=True 


            #step path
            if file_exists(self.steppath):
                self.steppath=stepfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".step"
                self.step=True 

            #threemf path
            if file_exists(self.threemfpath):
                self.threemfpath=threemffolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".threemf"
                self.threemf=True 

            #Model Edrawings path
            if file_exists(self.eprtpath):
                self.eprtpath=edrfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".eprt"
                self.eprt=True 
                self.edr=True
                self.edrpath=self.eprtpath
            elif file_exists(self.easmpath):
                self.easmpath=edrfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".easm"
                self.easm=True 
                self.edr=True
                self.edrpath=self.easmpath 

            #Drawing - Edrawings path
            if file_exists(self.edrwpath):
                self.edrwpath=edrfolder+self.file.replace(' ','%20')+"_REV_"+self.revision+".edrw"
                self.edrw=True
                self.edr_d=True
                self.edr_dpath=self.edrwpath

        self.get_process_icons()



    def get_process_icons (self):
        self.process_icons=[]
        self.process_colors=[]
        if self.process!="" and self.process in process_conf.keys() :
            self.process_icons.append('images/'+(process_conf[self.process]['icon']))
            self.process_colors.append(process_conf[self.process]['color'])
            if self.process2 in process_conf.keys() :
                self.process_icons.append('images/'+(process_conf[self.process2]['icon']))
                self.process_colors.append(process_conf[self.process2]['color'])
            if self.process3 in process_conf.keys() :
                self.process_icons.append('images/'+(process_conf[self.process3]['icon']))
                self.process_colors.append(process_conf[self.process3]['color'])
        else:
            self.process_icons.append('images/'+process_conf['others']['icon'])
            self.process_colors.append(process_conf['others']['color'])






    
    def __init__(self, partnumber="",revision="",description="",
                process="",process2="",process3="",
                finish="",path_to_model_file="",material="",matspec="",partType="",
                pdfpath="",edrpath="",edr_dpath="",jpgpath="",jpg_dpath="",pngpath="",png_dpath="",
                dxfpath="",dwgpath="",
                pdf="",edr="",edr_d="",jpg="",jpg_d="",png="",png_d="",dxf="",dwg="",check="",notes="",
                pdfindex="",pdfpages="",qty="",totalqty="",colour="",asset=""):

        self.partnumber=partnumber
        self.revision=revision
        self.description=description
        self.process=process
        self.process2=process2
        self.process3=process3
        self.finish=finish
        self.material=material
        self.colour=colour
        self.notes=notes
        self.asset=asset
       
      
        
        #outputfiles
        self.pdfpath=pdfpath
        self.pngpath=pngpath

                
        #Bom and index related
        self.qty=qty
        self.totalqty=totalqty
        self.pdfindex=pdfindex

        
    def ispartprocess(self,process):
       
        if process in self.process or process in self.process2 or process in self.process3:
            return True
        else:
            return False
   
 
    #To create a part from a dataframe row
    def partfromlist(self,datalist):
        self.spare_part=datalist['spare_part']
        self.mass=datalist['mass']
        self.thickness=datalist['thickness']
        
        self.category=datalist['category']
        self.datasheet=datalist['datasheet']
        
        self.partnumber=datalist['partnumber']
        self.process=datalist['process']
        self.process2=datalist['process2']
        self.process3=datalist['process3']
        self.configuration=datalist['configuration']
        self.file=datalist['file']
        self.folder=datalist['folder']
        self.supplier_partnumber=datalist['supplier_partnumber']
        self.description=datalist['description']
        self.finish=datalist['finish']
        self.material=datalist['material']
        self.revision=datalist['revision']
        self.approved=datalist['approved']
        self.author=datalist['author']
        self.supplier=datalist['supplier']
        self.link=datalist['link']
        self.oem=datalist['oem']
        self.treatment=datalist['treatment']
        self.drawndate=datalist['drawndate']
        self.colour=datalist['colour']
        #self.notes=datalist['notes']
        self.category=datalist['category']
        self.asset=datalist['asset']
                
        
        ##Extra values only needed for pdf list exports
        self.pdfpath=datalist['pdfpath']
        self.pngpath=datalist['pngpath']

        # try:
        #     self.qty=str(int(datalist[ 'qty']))
        # except:
        #     self.qty=0
        # self.totalqty=str(int(datalist[ 'totalqty']))
        try:
            self.qty=int(datalist[ 'qty'])
        except:
            self.qty=0
        try:
            self.totalqty=int(datalist[ 'totalqty'])
        except:
            self.totalqty=0
        
        self.png=False
        if os.path.isfile(self.pngpath): self.png=True
   
        
        try:
            self.pdfindex=str(int(datalist[ 'pdfindex']))
        except:
            self.pdfindex=""
        
        
        return self
   
    
    def get_components(self, components_only=True,process="",
                        flatbomtodict=False,
                        updatefiles=False, existing=[],existingqty=[],
                        list_consumed=False,consumedtodict=False):
        # print(existing,existingqty)
        
        # for part, qty in zip(existing,existingqty):
        #     print("########", part, qty)


        reflist=[]
        flatbom=[]
        



        try:
            if self.qty==None or self.qty==0 or self.qty=="":
                self.qty=1
            #     print("********** QTY OVERRIDEN *************")
            # else:
            #     print("** Father **",self.qty, self)
        except:
            self.qty=1
            # print("********** QTY OVERRIDEN *************")

        
    # if process=="receiving":
    #     for part in parts:

    #         if part.process is paint or welding or purchased or machined
    #             remove children

    #         if part.process has assembly only: ignoer


        def loopchildren(partnumber,revision,qty,reflist,process,list_consumed=False):
            part=Part.query.filter_by(partnumber=partnumber,revision=revision).first()
            children=part.children_with_qty()

            nochildrenlist=['paint','welding','purchase', 'machine']
            kids=True
            # #print("wtf")
            
            if process=="receiving":
                # #print(process)
                for processitem in nochildrenlist :
                    # #print(processitem)
                    if part.hasProcess(processitem):
                        kids=False
                        # #print("tonto",part) 
                        # if process=="welding":
                        #     #print("tonto",part) 
            
            if kids:
                for child in children:
                    refqty=child.qty*qty  
                    if part.hasConsumingProcess():
                        reflist.append([child,refqty,True])
                        # print('&&&&&&&&&&&&&&TRUE&&&&&&&&&&&&&&&&&&&&',child)
                    else:
                        reflist.append([child,refqty,False])
                        # print('===============FALSE===============',child)
                    # print("refqty",refqty,child)                  
                    if  len(child.children)>0:                        
                        # if not (child.hasConsumingProcess() and components_only):
                        #     loopchildren(child.partnumber,child.revision,refqty,reflist,process,list_consumed=list_consumed)
                        # elif child.hasConsumingProcess() and list_consumed:
                            loopchildren(child.partnumber,child.revision,refqty,reflist,process,list_consumed=list_consumed)

                    
        loopchildren(self.partnumber,self.revision,self.qty,reflist,process,list_consumed=list_consumed)
        
        #Sum up all quantities and compile flatbom
        resdict={}
        for item,q,consumed in reflist:

            total=resdict.get(item,{'total':0})['total']+q
            resdict[item]={'total':total,'consumed':consumed}
        
        for part in resdict.keys():
            part.qty=resdict[part]['total']
            addflatbom=True
            refqty=0
            for i in range(len( existing)):
                try:
                    item=existing[i]
                    
                    
                    if item['partnumber']==part.partnumber and item['revision']==part.revision:
                        # print("@@@@@@ existing @@@",item.qty, item)
                        # print("@@@@@@ existinglist @@@",existingqty[i], existing[i])
                        # print("@@@@@@ part     @@@",part.qty, part)
                        addflatbom=False
                        refindex=i
                        refqty=existingqty[i]
                except:
                    print("#####WTF####",print(type(item)), item['partnumber'])
                    print("#####WTF####",print(type(item)), item['partnumber'])
                    
            
            if addflatbom: 
                flatbom.append(part)                
            else:
                # print("@@@@@@berfore @@@", part.qty, part)
                part.qty=part.qty+refqty
                # print("@@@@@@REPEAETED@@@", part.qty, part)

        #Range flatbom by partnumber
        #flatbom.sort(key=lambda x: x.partnumber)
        # flatbom.sort(key=lambda x: (x.category,x.supplier,x.oem,x.approved,x.partnumber))
        
        #print(len(flatbom))   
        for part in flatbom:
            if updatefiles:
                part.updatefilespath(fileserver_path,local=True)
            
        
            for i in range(len( existing)):
                item=existing[i]
                # print("_______",item,part)
                
                
                
                if item['partnumber']==part.partnumber and item['revision']==part.revision:
                    # print("###### existing ###",item.qty, item)
                    # print("###### existinglist ###",existingqty[i], existing[i])
                    # print("###### part     ###",part.qty, part)

                    
                    qty_existing=existingqty[i]
                    
                    # print("###existing",item.qty, item)
                    # print("####part",part.qty, part)
                    # print("####refqty",resdict[part] )
                    
                    part.qty=part.qty+qty_existing

        cleanflatbom=[]
        consumed=[]

        for part in flatbom:
            if resdict[part]['consumed']:
                # print('@@@@@@@@consumedpart', part)
                consumed.append(part)
            else:
                cleanflatbom.append(part)
        
        # flatbom=cleanflatbom  
        # 
        if cleanflatbom!=[]:
            flatbom=cleanflatbom
            # print('@@@@@@@@clean',cleanflatbom)
            # print('@@@@@@@@consumed',consumed)

        
        if flatbomtodict:
            flatbomdict=[]
            for part in flatbom:
                dicto=part.as_dict()
                d2 = copy.deepcopy(dicto)                            
                d2['parent']=self.get_tag(addtimestamp=False)
                flatbomdict.append(d2)
            if flatbomdict!=[]:
                flatbom=flatbomdict


        # print('@@@@@@@@clean',cleanflatbom)
        # print('@@@@@@@@consumed',consumed)
        # print('@@@@@@@@flatbom',flatbom)
       
        if list_consumed:
            if consumedtodict:
                refconsumed=[]
                for conpart in consumed:
                    dicto=conpart.as_dict()
                    d2 = copy.deepcopy(dicto)                            
                    d2['parent']=self.get_tag(addtimestamp=False)
                    refconsumed.append(d2)

                # refconsumed=[x.as_dict() for x in consumed]
                return flatbom, refconsumed
            else:

                return flatbom, consumed
        else:
            return flatbom


    def partlist(bom_list):
        
        part_list=[]
        
        for i in range (len(bom_list)):
            part_in=Part()
            part_in=part_in.partfromlist(bom_list.iloc[i])
            part_list.append(part_in)
        return part_list
    

def get_tree(partnumber,revision,partlist,qty=1):
    
    refpart=Part.query.filter_by(revision=revision,partnumber=partnumber).first()
    ##print(partlist)    
    if len (refpart.children)>0: 
        for i in refpart.children:
            partlist.append((i.getchild(),i.qty*qty))
            get_tree(i.child.partnumber,i.child.revision,partlist,qty=i.qty*qty)
        
    return partlist



def get_flatbom(partnumber,revision,qty=1):
    refpart=Part.query.filter_by(revision=revision,partnumber=partnumber).first()
    flatlist=[]
    flatlist.append((refpart,qty))
    
    dictlist=[]
    checklist=[]
    
    
    get_tree(partnumber,revision,flatlist)
    
    for part,part_qty in flatlist:
        part_dict={}
        part_dict=part.__dict__.copy()
        # part_dict.pop('_sa_instance_state')
        
        if part_dict['partnumber']+part_dict['revision'] in checklist:
            ref_index=checklist.index(part_dict['partnumber']+part_dict['revision'])
            dictlist[ref_index]['totalqty']+=part_qty 
        else:
            part_dict['totalqty']=part_qty  
            dictlist.append(part_dict)
            checklist.append(part_dict['partnumber']+part_dict['revision'])
        
    ##print(dictlist)
    
    flatbom=pd.DataFrame(dictlist)
    flatbom['pdfpath']=deliverables_folder+"pdf\\"+flatbom['file']+"_REV_"+flatbom['revision']+".pdf"
    flatbom['pngpath']=deliverables_folder+"png\\"+flatbom['file']+"_REV_"+flatbom['revision']+".png"
    
        
    
    return flatbom
    



