data=test.data
flatbom=test.flatbom
bom=test.bom

top_partnumber=test.partnumber
top_rev=test.revision


import win32com.client
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import openpyxl
import os




def excel_to_pdf (excel_path, pdf_path,erasepdf=False):

    o = win32com.client.Dispatch("Excel.Application")
    
    o.Visible = False
    
    wb = o.Workbooks.Open(excel_path)
    
    # ws_index_list = [1] #say you want to print these sheets
    
    # ws=o.ActiveSheet
    # wb.WorkSheets(ws).Select()
    
    wb.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
    wb.Close(SaveChanges=False)
    if erasepdf:
        os.remove(excel_path)





def get_files(bom_in,extension,partfolder):
     
    outputfolder=partfolder+"/"+extension+"/"

       
    for index, row in bom_in.iterrows():
        filenamebit=row["partnumber"]+"_REV_"+row["revision"]
        sourcefile=deliverables_folder+extension+"/"+filenamebit+"."+extension
        
        if extension=='dxf':
            targetfile=outputfolder+filenamebit+"-"+row["material"]+"_"+str(row["thickness"])+"mm."+extension
            # if row["thickness"]==0 or str(row["thickness"])=='':
            #    bom_in.at[index,'notes'].append("Missing thickness")
        else:
            targetfile=outputfolder+filenamebit+"."+extension
            #print(targetfile)
        
        try:
            if os.path.exists(sourcefile):
                create_folder_ifnotexists(outputfolder)
            copyfile(sourcefile,targetfile)
        except:
            #print("coouldnt get file ", extension," for ",row['partnumber'])
        

def get_man_bom(top_partnumber,top_rev,bom, flatbom):
    
    man_list=[]
    def get_man_children(partnumber,revision,bom, flatbom):
        children=get_children(partnumber,revision,bom, flatbom)
        for index,row in children.iterrows():
            if row['process']=='welding' or row['process2']=='welding' or row['process3']=='welding':
                man_list.append(row)
            else:
                try:
                    get_man_children(row['partnumber'],row['revision'],bom,flatbom)
                except:
                    pass
    get_man_children(top_partnumber,top_rev,bom, flatbom)
    manu_df=flatbom.drop(flatbom.index)

    for row in man_list:
        manu_df=manu_df.append(row)
        
    manu_df=manu_df.drop_duplicates(subset=['partnumber', 'revision']).reset_index(drop=True)
    
    return manu_df




def get_flat_bom( partnumber,revision,bom, flatbom,qty=1):
    
    children_list=[]
    def get_recursive(partnumber,revision,bom, flatbom,qty):
        children=get_children(partnumber,revision,bom, flatbom)
        if len(children)==0: 
            return None
        else:
            
            for index,row in children.iterrows():
                    rowmod=row.copy()
                    rowmod.at['qty']=row.at['qty']*qty
                    children_list.append(rowmod)
                    try:
                        get_recursive(row['partnumber'],row['revision'],bom,flatbom,row['qty'])
                    except:
                        pass
                    
    get_recursive(partnumber,revision,bom, flatbom,qty)
    
    if len (children_list)==0:
        return flatbom.drop(flatbom.index)
    
    
    children_df=flatbom.drop(flatbom.index)
    for row in children_list:
        children_df=children_df.append(row)
        
    children_qty=children_df[['partnumber','revision','qty']].groupby(by=['partnumber', 'revision'], axis=0).sum()
    
    children_df=children_df.drop_duplicates(subset=['partnumber', 'revision']).reset_index(drop=True)
    
    for (partnumber,revision),row in children_qty.iterrows():
       
        ref_index=children_df.loc[(children_df['partnumber']==partnumber) & (children_df['revision']==revision)].index.values[0]
        children_df.at[ref_index,'qty']=row['qty']
    
 
    return children_df


def bom_to_excel(bom_in,outputfolder,title="",qty="qty"):
    
        sheet="Manufactured components"
 
        if title=="": title="List"
        
        excel_file=outputfolder + title +".xlsx"
        
        property_list=['partnumber','revision',qty,'description','material','thickness','finish']
        bom_to_sheet=bom_in[property_list] 
        
        
        #Generate basic lists:
        with pd.ExcelWriter(excel_file) as writer:
            bom_to_sheet.to_excel(writer, sheet_name=sheet)
            
            


 
        
        workbook=load_workbook(filename=excel_file)
        # # #print(dir(self.workbook))
        
        # Add images
        worksheet=workbook[sheet]
        worksheet.insert_cols(idx=2)
        

        
        #Get the a
        str_flatbom=bom_to_sheet.applymap(str)
        col_width=[]

        for col in str_flatbom:
           col_width.append(str_flatbom[col].map(len).max())
        col_width.insert(0,5)
        col_width.insert(1,10)

        
        
        #print(col_width)
       
        for i, column_width in enumerate(col_width):
            if column_width>5:
                worksheet.column_dimensions[get_column_letter(i+1)].width =int( 1.1*column_width)


        
                

        i=-1           
        for index, row in bom_in.iterrows():
            i=i+1
            rd =worksheet.row_dimensions[i+2] # get dimension for row 3
            rd.height = 40 # value in points, there is no "auto"
            png_folder=deliverables_folder+"png/"
            thumb=thumbnail(png_folder+row.at['file']+"_REV_"+row['revision']+'.png',
                            size=(100, 100))
            

            for j,col in enumerate(maincols):
        
                
                if row[col] =='' or str(row[col])=='nan':
                    cell=chr(67+j)+str(i+2)
                    if col!="process2" and col!="process3" : 
                        worksheet[cell].fill=redFill
 
                    
            #Add revision color warning
            if row['revision'] =='':
                worksheet['D'+str(i+2)].fill = yellowFill
            # #Add description color warning
            # if self.flatbom.at[i,'description'] =='' or str(self.flatbom.at[i,'description'])=='nan':
            #     worksheet['E'+str(i+2)].fill = redFill
            
            
            
            #Add image
            cell='B'+str(i+2)
            try:
                image=openpyxlIm(thumb)
                image.height=60
                image.width=60
                worksheet.add_image(image, cell)
                
            except:
                # worksheet.add_image(thumb, cell)
                worksheet[cell].fill = redFill
                #print("Could not add image to excel ", row['partnumber'])
                
            
        
        #create the border style to put all around the cells
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))   

        #Border annd Center alignment for all cells    
        for row in worksheet.iter_rows():
            for cell in row:
                # #print(dir(cell.style))
                # cell.style.alignment.wrap_text=True
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        #Put to landscape and adjust the width of the page to the width of the content
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(worksheet, paper_size = int(sum(col_width)), orientation='landscape')
        workbook.save(filename=excel_file)
        
        #Export to pdf
        pdf_path=outputfolder + title +".pdf"
        excel_to_pdf (excel_file, pdf_path,erasepdf=True)
        
        return excel_file





testchildren=get_flat_bom(test.partnumber,test.revision,bom, flatbom) [['partnumber','revision','qty']]


manbom=get_man_bom(test.partnumber,test.revision,bom, flatbom)

weld_self=[]
flat_list=[]


#Check if outputfolder exists otherwise create it

create_folder_ifnotexists(test.folderout+"manufacturing/")

bomtitle=test.tag+"-manufacturing list-"

bom_to_excel(manbom,test.folderout+"manufacturing/",title=bomtitle,qty='totalqty')


for index, row in manbom.iterrows():
    
        flat=get_flat_bom(row['partnumber'],row['revision'],bom,flatbom)
        
        
        #Create folder and copy the dxf files      
        outputfolder=test.folderout+"manufacturing/"+row['partnumber']+"/"
        create_folder_ifnotexists(outputfolder)
        
        if len(flat)>0:
            #To add the top level to the list at the begining
            oneflat=manbom.drop(manbom.index)
            oneflat=oneflat.append(row)
            
            #Only for the excel list
            flat_excel=flat.copy()
            flat_excel=flat_excel.loc[(flat_excel['process']!='welding') & (flat_excel['process2']!='welding') & (flat_excel['process3']!='welding')]
            flat_excel=pd.concat([oneflat, flat_excel], ignore_index=True).reset_index(drop=True)
            
            
            get_files(flat,'dxf',outputfolder)
            flat=pd.concat([oneflat, flat], ignore_index=True).reset_index(drop=True)
            
            
            
            # get_files(flat,'pdf',outputfolder)
            # bom_to_excel(flat,outputfolder,title=row['partnumber']+'_REV_'+row['revision'])
        else:
            flat=manbom.drop(manbom.index)
            flat=flat.append(row).reset_index(drop=True)
            flat_excel=flat.copy()
            
        get_files(flat,'dxf',outputfolder)
        # get_files(flat,'pdf',outputfolder)
        bom_to_excel(flat_excel,outputfolder,title=row['partnumber']+'_REV_'+row['revision'] + "-BOM")
            
            
        fakebom=solidbom("",deliverables_folder,outputfolder)
        fakebom.partnumber=row['partnumber']
        fakebom.revision=row['revision']
        fakebom.description=row['description']
        fakebom.tag=fakebom.partnumber+"_REV_"+fakebom.revision+"-"+fakebom.timestamp
        fakebom.folderout=outputfolder
        fakebom.flatbom=flat
        IndexPDF(fakebom,outputfolder=outputfolder,sort=False)

    