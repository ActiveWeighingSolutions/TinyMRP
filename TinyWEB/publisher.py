# boxy_line_demo.py
# from TinyMRP import *
# from TinyMRP_PDFlibraries import *
# from TinyMRP_templates import BoxyGrid

# from models import Part
from TinyWEB.models import Part, solidbom, thumbnail, file_exists, create_folder_ifnotexists, qr_code
from TinyWEB import process_conf, variables_conf, webfileserver, deliverables_folder, fileserver_path, deliverables, db
from TinyWEB import maincols, refcols, webserver, app
import requests, io


#print("loaded PUBLISHER")

#For excel handling
# import win32com.client
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as openpyxlIm
import openpyxl

#QR code library
import qrcode

#ReportLab libraries
from reportlab.lib.pagesizes import A4, A3, landscape
from reportlab.platypus import SimpleDocTemplate, Spacer, Frame, PageTemplate
from reportlab.lib.units import inch, mm
from reportlab.rl_config import defaultPageSize
from reportlab.platypus import BaseDocTemplate, Frame, Paragraph, NextPageTemplate, PageBreak, PageTemplate
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.rl_config import defaultPageSize
from reportlab.platypus import Image, Frame, PageTemplate, BaseDocTemplate
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

#For adding the svg
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPM


# boxy_line.py
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Flowable, SimpleDocTemplate, Spacer
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.graphics.shapes import Rect, Ellipse, Circle

#Form libraries
from reportlab.pdfbase import pdfform
from reportlab.lib.colors import magenta, pink, blue, green
from reportlab.lib.units import inch, mm
from reportlab.rl_config import defaultPageSize

from reportlab.graphics import renderPDF

from svglib.svglib import svg2rlg


#For crating the thumbnail
import os, sys
from PIL import Image

#for moving and copying files
from shutil import copyfile

#Styles libraries
from reportlab.lib.styles import ParagraphStyle


#Form libraries
from reportlab.pdfbase import pdfform
from reportlab.lib.colors import magenta, pink, blue, green


from reportlab.platypus import Preformatted,  XPreformatted
from reportlab.platypus.tableofcontents import SimpleIndex

#Database libraries
from flask_sqlalchemy import SQLAlchemy

from flask import (
    Blueprint, flash, g, redirect, session, render_template, request, url_for
)


#File handling libraries
import os
import fnmatch
from pathlib import Path
from datetime import datetime, date #for timestamps

#PDF libraries
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
import pdfkit #TO EXPORT WEBPAGES TO PDF





#REad bom txt/cvs file
import pandas as pd
import numpy as np
import re
import glob
from pathlib import Path


#Reportlab tables
from reportlab.platypus import Paragraph, Frame, Table, Spacer, TableStyle

#Database libs
from sqlalchemy import select, or_, and_



print(app.config['PROCESS_LEGEND'])




# Create fill fore excel boms warning
redFill = PatternFill(start_color='EE1111',
               end_color='EE1111',
               fill_type='solid')

yellowFill= PatternFill(start_color='FFFF00',
               end_color='FFFF00',
               fill_type='solid')


def makeParaCenter(text):
    output="<para align=center>"
    output=output+text
    output=output+"</para>"
    
    return output
    
def makeParaRight(text):
    output="<para align=right>"
    output=output+text
    output=output+"</para>"
    
    return output







# def thumbnail(infile):

#     size = 400, 400

#     outfile = os.path.splitext(infile)[0] + ".thumbnail.png"
#     if infile != outfile:
#         try:
#             im = Image.open(infile)
#             im.thumbnail(size, Image.ANTIALIAS)
#             im.save(outfile, "PNG")
#             return outfile
#         except IOError:
#             print ("cannot create thumbnail for '%s'" % infile)


def index_page(canvas,doc):
    
    logo="TinyWEB/static/images/"+'logo.png'
    PAGE_HEIGHT = A4[1]
    PAGE_WIDTH = A4[0]
 
    canvas.setFillColor(colors.white)
    # canvas.rect(0, PAGE_HEIGHT-90, PAGE_WIDTH, 90, stroke=1, fill=1)
    canvas.drawImage(logo,inch*0.5-PAGE_WIDTH, PAGE_HEIGHT-50,height=0.5*inch, 
                     preserveAspectRatio=True, mask='auto')

    canvas.saveState()





    
def process_page(process):
    
    icon=process_conf[process]['icon']
    drawing = svg2rlg("TinyWEB/static/images/"+icon)
    drawing.scale(0.1,0.1)
    
    
    def process_genericpage(canvas,doc):
        #inputs process, process icon, and 
        logo="TinyWEB/static/images/"+'logo.png'
        PAGE_HEIGHT = A4[1]
        PAGE_WIDTH = A4[0]
     
        
        
        # canvas.setStrokeColor(get_process_color(process))
        canvas.setFillColor(get_process_color(process,intensity=0.7))
        canvas.rect(0, 0, PAGE_WIDTH, PAGE_HEIGHT, stroke=0, fill=1)
        
        canvas.setFillColor(colors.white)
        canvas.rect(20, 20, PAGE_WIDTH-40, PAGE_HEIGHT-40, stroke=0, fill=1)
        
        
        #Inser company logo and process icon
        canvas.drawImage(logo,inch*0.5-PAGE_WIDTH, PAGE_HEIGHT-70,height=0.5*inch, 
                         preserveAspectRatio=True, mask='auto')
        
        renderPDF.draw(drawing, canvas,PAGE_WIDTH-inch*1.5, PAGE_HEIGHT-80)
    
        canvas.saveState()
    
    return process_genericpage
    



def get_process_color(process, intensity=1):
    #Function to provide the reportlab color format from the configuration
    try:
        color=process_conf[process]['color'].split(",")
    except:
        color=process_conf['others']['color'].split(",")
    
    color=[float(col)/255.0 for col in color ]
 
        
    return colors.Color(color[0],color[1],color[2],intensity)
# Add foot note to pdf page object
def pdf_pagenum(pdf_page,page_number,total="",color=colors.grey):

    if total=="":
        footnote= str(page_number+1) 
    else:
        footnote= str(page_number+1) +"/"+str(total)
    
    #Location of the page number
    x=30
    y=25
    w=30
    h=20
    yvert=y

    
    #OPen page and get properties
    PAGE_WIDTH=int(pdf_page.mediaBox.getWidth())
    PAGE_HEIGHT=int(pdf_page.mediaBox.getHeight())

    #Create temp page to merge with doc page
    tempfile="temp/temp_page_"+ str(int(datetime.now().timestamp()*1e6))+".pdf"
    c=canvas.Canvas(tempfile)




    #Draw rectangle and foot depending on orientation
    if PAGE_WIDTH<PAGE_HEIGHT:
                    
        # c.rotate(90)
        
        # c.setStrokeColorRGB(0,0.8,0)
        c.setFillColor(color)
        c.rect(PAGE_WIDTH-x-w/2, yvert-h/2, w, h, stroke=0, fill=1)
        
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(colors.white)
        c.drawCentredString(PAGE_WIDTH-x,yvert-2*mm,footnote)
       
    else:
        
        # c.setStrokeColorRGB(0,0.8,0)
        c.setFillColor(color)
        c.rect(PAGE_WIDTH-x-w/2, y-h/2, w, h, stroke=0, fill=1)
        
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(colors.white)
        c.drawCentredString(PAGE_WIDTH-x,y-2*mm,footnote)
      
    #Save page
    c.showPage()
    c.save()
    

    #Merge the pdf page with watermark
    watermark = PdfFileReader(tempfile)
    watermark_page = watermark.getPage(0)
    pdf_page.mergePage(watermark_page)
    
        
    #Remove temporary file
    os.remove(tempfile)

    return pdf_page    



def IndexPDF(solidbom,outputfolder="",savevisual=False,sort=True,title="",norefpart=False,showvisual=True):
    
    if outputfolder=="":
        outputfolder=solidbom.folderout

    if title=="" and not norefpart:
        output_path=outputfolder +  solidbom.partnumber+ "_REV_"  + solidbom.revision+ \
                "-"+ solidbom.description+"-Drawing_Pack-"+ ".pdf"
    elif norefpart:
        output_path=outputfolder +  "Drawing_Pack-" +str(int(datetime.now().timestamp()*1e6))+".pdf"
    else:
        output_path=outputfolder +  title + ".pdf"

    tempPath="temp/temppack_"+ str(int(datetime.now().timestamp()*1e6))+".pdf"
    
    
    data2=solidbom.flatbom
    
    if sort :
        data2=data2.sort_values('partnumber', axis=0, ascending=True, inplace=False, kind='quicksort', na_position='last').reset_index()


    #print('INSIDE INDEX')
    # for i,pepe in data2.iterrows():
    #     print (pepe.partnumber)
    
    data2["pdfpages"]=0
    
    
    
    #Set the first page and genertate the wrtier object
    pageref=1
    pdfWriter = PyPDF2.PdfFileWriter()
    
    
    #Index page code generation
    dirtyindexfile="dirtyindex.pdf"
    doc = SimpleDocTemplate(dirtyindexfile,    pagesize=A4,
                        rightMargin=40,      leftMargin=50,
                        topMargin=80,      bottomMargin=20)
    styles = getSampleStyleSheet()
    #Modify the normal style to be smaller font
    styles['Normal'].fontSize = 10
    styles['Heading1'].textColor = colors.black
    flowables = []
     
    #Index page template reference
    index_firsframe = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height+60, id='normal')
    index_nextframe = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    doc.addPageTemplates([PageTemplate(id='index_next',frames=index_nextframe,pagesize=A4,onPage=index_page),
                         PageTemplate(id='index_first',frames=index_firsframe,pagesize=A4,onPage=index_page),])
    


    
    
    
    #Compile all the files
    
    for i in range(len(data2)):
        ##print(data2.at[i,"pdfpath"])
        if os.path.isfile(data2.at[i,"pdfpath"]):# or file_exists("http://"+data2.at[i,"pdfpath"]):
                
                pdfFileObj = open(data2.at[i,"pdfpath"],'rb')
                pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
            
                #Bookmark and index string
                bookmark=data2.at[i,'partnumber'] + " Rev: " + data2.at[i,'revision'] + " - "+ data2.at[i,'description']
                bookmark=bookmark.replace(',',',,')
            
                #Add the page index ref and the amount of pages to data2    
                data2.at[i,"pdfindex"]=pageref
                data2.at[i,"pdfpages"]=pdfReader.numPages
            
                ptext = """ I'm a phantom page<index item=" """
                ptext=ptext+bookmark.replace(',,',';')
                ptext=ptext+""" "/>bulletted paragraph"""
                para = Paragraph(ptext, style=styles["Normal"], bulletText='-')
                flowables.append(para)
            
                #Opening each page of the PDF
                for pageNum in range(data2.at[i,"pdfpages"]):
                    pageObj = pdfReader.getPage(pageNum)
                                    
                    pageObj=pdf_pagenum(pageObj,pageref+pageNum-1,color=get_process_color(data2.at[i,"process"]))
                
                    # pageObj=pdf_pagenum(pageObj,pageref+pageNum-1,color=colors.black)
                    pdfWriter.addPage(pageObj)
                
                    #Add the phantom pages
                    flowables.append(NextPageTemplate('index_first'))
                    flowables.append(PageBreak())
                
                            
                #Add bookmark
                try:
                    pdfWriter.addBookmark(bookmark,pageref-1)
                except:
                    pass
                    #print("couldnt add bookmark for ",data2.at[i,'partnumber'])
                pageref=pageref+data2.at[i,"pdfpages"]
        
        else:
            data2.at[i,"pdfpath"]=""
            #print("No file for ",data2.at[i,"partnumber"])
    
    #Store back the index ref into the solidbom
    solidbom.flatbom=data2
    

    
    with open(tempPath, 'wb') as fh:
        pdfWriter.write(fh)
    
    
    #Build the index
    index = SimpleIndex(dot='.', headers=False)
    
    
    # headertext=makeParaRight(solidbom.partnumber + " Drawing Pack")
    headertext=makeParaCenter("DRAWING PACK INDEX" )
    para = Paragraph(headertext, style=styles["Heading1"])
    flowables.append(para)

    if not norefpart:
        headertext=makeParaCenter(solidbom.partnumber+  " REV "+solidbom.revision + ": "+solidbom.description)
        para = Paragraph(headertext, style=styles["Heading2"])
        flowables.append(para)
    
    
    flowables.append (Spacer(0, 20))
    flowables.append(NextPageTemplate('index_next'))
    flowables.append(index)

    # print(data2.loc[0,:])
    # print(data2.loc[1,:])
    
    doc.build(flowables, onFirstPage=index_page, onLaterPages=index_page,canvasmaker=index.getCanvasMaker())
    
    #Create the visual index to add it to the drawing pack
    if showvisual or savevisual:
        visualindexfile=visual_list(solidbom, outputfolder=outputfolder,norefpart=norefpart)
        visualindex=PdfFileReader(visualindexfile)
    


    
    splitter=PyPDF2.PdfFileWriter()
    dirtyindex=PdfFileReader(dirtyindexfile)
    cleanindexfile="clean_index.pdf"
    
    #Extract the index
    j=0
    for page in range(pageref-1,dirtyindex.getNumPages()):
        splitter.addPage(dirtyindex.getPage(page))
        j=j+1
        
     #Add index bookmark
    splitter.addBookmark("Index",0)   
        
    #If visual index need add to page
    if showvisual:
        for page in range(0,visualindex.getNumPages()):
            splitter.addPage(visualindex.getPage(page))
            
        #Add vsiual index bookmark
        splitter.addBookmark("Visual index",j)   
    
    with open(cleanindexfile, 'wb') as fh:
        splitter.write(fh)
 
    
    
    # #Merge compiledpdfs and index
    pdf_merger = PdfFileMerger()
    pdf_merger.append(cleanindexfile)
    pdf_merger.append(tempPath)

    
    
    with open(output_path, 'wb') as fileobj:
        pdf_merger.write(output_path)
    pdf_merger.close()
    

    
    #Remove temp files
    os.remove(dirtyindexfile)
    os.remove(tempPath)
    os.remove(cleanindexfile)
    
    if showvisual or savevisual:
        os.remove(visualindexfile)
        
    
    
    return output_path

def pdfListCompile(listofpdf,output_path):
    pdf_merger = PdfFileMerger()

    for pdffile in listofpdf:
        pdf_merger.append(pdffile)
    
    with open(output_path, 'wb') as fileobj:
        pdf_merger.write(output_path)
    pdf_merger.close()
    



def visual_list(solidbom, outputfolder="",title="",norefpart=False):
    

    #get children data
    bom_in=solidbom.flatbom

    #Filename generation
    if outputfolder=="":
        outputfolder=solidbom.folderout

    if title=="" and not norefpart:
        outputfile=outputfolder + "Visual_list-"+ solidbom.tag+".pdf"
    elif norefpart:
        outputfile=outputfolder +  "Visual_list-"+str(int(datetime.now().timestamp()*1e6))+".pdf"
    else:
        outputfile=outputfolder + title + "-"+ solidbom.tag+".pdf"




    #In case the list is empty
    if len(bom_in)==0:
        return
    

    #Tempfile
    tempfile="temp/temp_visual_"+ str(int(datetime.now().timestamp()*1e6))+".pdf"

    doc = SimpleDocTemplate(tempfile,    pagesize=A4,
                            rightMargin=20,      leftMargin=30,
                            topMargin=20,      bottomMargin=20)
    pagebreak = PageBreak()
    styles = getSampleStyleSheet()
    
    #Sections building

    
    #Index page template reference
    index_firsframe = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    index_nextframe = Frame(doc.leftMargin, doc.bottomMargin+20, doc.width, doc.height-60, id='normal')
    doc.addPageTemplates([PageTemplate(id='index_first',frames=index_firsframe,pagesize=A4,onPage=index_page),
                         PageTemplate(id='index_next',frames=index_nextframe,pagesize=A4,onPage=index_page),])


    #Create the story for the pdf
    flowables=[]
    
    #Add pdf header
        # headertext=makeParaRight(solidbom.partnumber + " Drawing Pack")
    
    if title=="":
        headertext=makeParaCenter("Visual index" )
    else:
        headertext=makeParaCenter(title)
    para = Paragraph(headertext, style=styles["Heading1"])
    flowables.append(para)
 
    if not norefpart:
        headertext=makeParaCenter(solidbom.partnumber+  " REV "+solidbom.revision + ": "+solidbom.description)
        para = Paragraph(headertext, style=styles["Heading2"])
        flowables.append(para)
    
    
    flowables.append (Spacer(0, 20))
    flowables.append(NextPageTemplate('index_next'))


            
    partlist=Part.partlist(bom_in)

    nofastener=[x for x in partlist if not x.hasProcess('hardware')]
    fasteners=[x for x in partlist if  x.hasProcess('hardware')]
    
    #Split in sections
    boxywidth=60
    boxyheight=45
    boxycols=int(((210-20)/ boxywidth))
    margin=30
    spacer=2
    boxyrows=int(((297-margin)/ (boxyheight+spacer))-1)
    

    chunks = [nofastener[x:x+boxycols] for x in range(0, len(nofastener), boxycols)]
    #flash(boxycols)
   
    #Add all the parts that are not fasteners
    j=0
    firstrow=False
    flowables.append (Spacer(0, boxyheight/2*mm))
    for split_list in chunks:
        if firstrow:
            flowables.append(pagebreak)
            flowables.append (Spacer(0, boxyheight/2*mm))
            firstrow=False
            
        grid=BoxyGrid(split_list, width=boxywidth, height=boxyheight,y=-j*boxyheight/2)
        j=j+1
        flowables.append(grid)
        flowables.append (Spacer(0, spacer*mm))
        if j>boxyrows+1:
            flowables.append(NextPageTemplate('index_next'))
            j=0
            firstrow=True

    flowables.append (Spacer(0, boxyheight*mm))

    
    
    #Add process legend
    if j>4 or firstrow:
        flowables.append(pagebreak)
        # flowables.append (Spacer(0, 60*mm))
    
    flowables.append (Spacer(0, boxyheight*mm))
        
    flowables.append(processlengend(app.config['PROCESS_LEGEND']))  
    flowables.append (Spacer(0, boxyheight*mm))
    
    #Add the fasteners table:
    if len(fasteners)>0:
        headertext=makeParaCenter("Fasteners list")
        para = Paragraph(headertext, style=styles["Heading1"])
        
        # flowables.append(pagebreak)
        flowables.append(para)
        flowables.append(partTable(fasteners))  
    

    doc.build(flowables,onFirstPage=index_page, onLaterPages=index_page )


    #Add page numbers
    addPageNumbersOnly (tempfile,outputfile)
    os.remove(tempfile)





    return outputfile


   
        
        
class BoxyGrid(Flowable):
    """
    Draw a box + line + text
    -----------------------------------------
    | foobar |
    ---------
    """
    def __init__(self, children, x=0,y=0,width=45.4, height=45.4):
        Flowable.__init__(self)
        self.x = x
        self.y = y
        self.width = width
        self.height = height
        # self.title = str(part.partnumber) + " - Rev " + str(part.rev)
        # self.qty=qty
        # self.part=part
        self.children=children
        
        
    def draw(self):
        xrows=0
        yrows=0
        children=self.children

        #Place image
        x=self.x*mm
        y=self.y*mm
        
        xc=self.width*mm
        yc=self.height*mm
       
        
        row=0
        j=0
        
        for i in range(len(children)):
            
            if j>=int((A4[0]-inch*1)/xc):
                row=row+10
                j=0
                
            child=children[i]
            xlevel=x+j*xc
            ylevel=y-row*yc
            qrsize=min([yc*0.9,xc*0.9,50])
            

            if child.png:
                # BoxyLine(part=child, x=x,y=ylevel)
                
                self.canv.drawImage(thumbnail(child.pngpath),xlevel+10, ylevel, height=yc*0.9, preserveAspectRatio=True, mask='auto')
                
            if self.width>45 and self.height>25:
                qrimagepath=qr_code(child)
                self.canv.drawImage(qrimagepath,xlevel+xc-qrsize, ylevel+yc-qrsize-yc*0.21, width=qrsize,height=qrsize)
            if child.approved:
                # icon=svg2rlg("TinyWEB/static/images/"+'approved.svg')
                # icon=renderPM.drawToFile(icon, "TinyWEB/static/images/"+'approved.png', fmt='PNG')
                icon="TinyWEB/static/images/"+'approved.png'
                self.canv.drawImage(icon,xlevel+xc-qrsize, ylevel+yc-qrsize-yc*0.21-20, width=30,height=20)
            # self.canv.drawImage(qr_code(child),xlevel+20, ylevel+30,width=xc, height=xc)
                
            self.canv.setFont("Helvetica-Bold", 10)
            self.canv.setFillColor(colors.darkblue)
             
            #Draw rectangle
            fill=0
            self.canv.setLineJoin(1)
            
            self.canv.setStrokeColor(get_process_color(child.process))
            self.canv.setLineWidth(3)
            self.canv.rect(xlevel+xc*0.1, ylevel+0.05*yc, xc*.92, yc*.82,fill=fill)
            self.canv.setLineWidth(2)
   
            if child.process2!="":
                    indent=0.03
                    self.canv.setStrokeColor(get_process_color(child.process2,intensity=0.7))
                    self.canv.rect(xlevel+xc*(0.1+indent), ylevel+(0.05+indent)*yc, xc*(0.92-2*indent), yc*(0.82-2*indent),fill=fill)
            if child.process3!="":
                    indent=0.06
                    self.canv.setStrokeColor(get_process_color(child.process3,intensity=0.7))
                    self.canv.rect(xlevel+xc*(0.1+indent), ylevel+(0.05+indent)*yc, xc*(0.92-2*indent), yc*(0.82-2*indent),fill=fill)
            self.canv.setLineWidth(1)    
             
            ytop=ylevel+yc*0.8-1.52*mm     
            # self.canv.drawCentredString(xlevel+xc/2, ylevel+xc*0.5, child.partnumber)
            #self.canv.drawString(xlevel+xc*0.1 + 1.5*mm, ytop, child.partnumber)
            

            #Part number text
            if len(child.partnumber)>14:
                chunks=[child.partnumber[x:x+14] for x in range(0, len(child.partnumber.upper()), 14)]
                k=0
                for chunk in chunks:
                
                    self.canv.drawString(xlevel+xc*0.1 + 1.5*mm, ytop-k, chunk)
                    k=k+10
                ytopleft=ytop-k
            else:
                self.canv.drawString(xlevel+xc*0.1 + 1.5*mm, ytop, child.partnumber)
                ytopleft=ytop-10

                           
            
            
            if child.revision:
                self.canv.setFont("Helvetica", 8)
                self.canv.setFillColor(colors.grey)
                self.canv.drawString(xlevel+xc*0.1 + 1.5*mm,ytopleft , "REV: " + child.revision)
                ytopleft=ytopleft-10
            
            #Page index mark
            if child.pdfindex!="":
                self.canv.setFont("Helvetica", 8)
                self.canv.setFillColor(colors.grey)
                self.canv.drawString(xlevel+xc*0.1 + 1.5*mm, ytopleft, "P: " +child.pdfindex)
            
            #Qty mark
            self.canv.setFont("Helvetica-BoldOblique", 12)
            self.canv.setFillColor(colors.darkred)
            if child.qty!=0 and child.qty!=child.totalqty:
                self.canv.drawRightString(xlevel+xc-1*mm,ytop , "x" + str(child.qty))
                self.canv.setFont("Helvetica", 8)
                self.canv.setFillColor(colors.grey)
                self.canv.drawRightString(xlevel+xc-1*mm,ytop-10 , "of " + str(child.totalqty))
            else:
                self.canv.drawRightString(xlevel+xc-1*mm,ytop , "x" + str(child.totalqty))
                

            
            self.canv.setFont("Helvetica-Bold", 8)
            self.canv.setFillColor(colors.black)
            
            textlimit=int(xc*22/(45*mm))
            chunks=[child.description[x:x+textlimit] for x in range(0, len(child.description.upper()), textlimit)]
            
            # if len(chunks)==1:
            #     k=-10
            # elif len(chunks)==2:
            #     k=0
            # elif len(chunks)==3:
            #     k=10
            k=(1-len(chunks))*10+10
            for chunk in chunks:
                self.canv.drawString(xlevel+xc*0.1 + 1.5*mm, ylevel+yc*0.05+13-k, chunk)
                k=k+10               
                
               
            #next child    
            j=j+1
            
            

                    

            


def process_visual_list(solidbom,process):
    
    #get children data
    bom_in=solidbom.flatbom
    bom_in=bom_in.loc[(bom_in['process']==process) | (bom_in['process2']==process) | (bom_in['process3']==process)]
    
    outputfile=solidbom.folderout + process +" components for "+ solidbom.tag+".pdf"
 
    #In case the list is empty
    if len(bom_in)==0:
        return
    
    doc = SimpleDocTemplate(outputfile,    pagesize=A4,
                            rightMargin=20,      leftMargin=30,
                            topMargin=20,      bottomMargin=20)
    pagebreak = PageBreak()
    styles = getSampleStyleSheet()
    
    #Sections building

    
    #Index page template reference
    index_firsframe = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height-25, id='normal')
    index_nextframe = Frame(doc.leftMargin, doc.bottomMargin+20, doc.width, doc.height-80, id='normal')
    doc.addPageTemplates([PageTemplate(id='index_first',frames=index_firsframe,pagesize=A4,onPage=process_page(process)),
                         PageTemplate(id='index_next',frames=index_nextframe,pagesize=A4,onPage=process_page(process)),])


    #Create the story for the pdf
    flowables=[]
    
    #Add pdf header
        # headertext=makeParaRight(solidbom.partnumber + " Drawing Pack")
    headertext=makeParaCenter(process.upper() + " components for" )
    para = Paragraph(headertext, style=styles["Heading1"])
    flowables.append(para)
    headertext=makeParaCenter(solidbom.partnumber+  " REV "+solidbom.revision + ": "+solidbom.description)
    para = Paragraph(headertext, style=styles["Heading2"])
    flowables.append(para)
    
    
    flowables.append (Spacer(0, 0.1))
    flowables.append(NextPageTemplate('index_next'))


            
    partlist=Part.partlist(bom_in)
    
    #Split in sections
    chunks = [partlist[x:x+4] for x in range(0, len(partlist), 4)]
   
    #Add all the parts
    j=0
    for split_list in chunks:
        grid=BoxyGrid(split_list, width=45*mm,y=-50)
        flowables.append(grid)
        flowables.append (Spacer(0,15*mm))
        j=j+1
        if j>5:
            flowables.append(NextPageTemplate('index_next'))
            flowables.append(pagebreak)
            flowables.append(NextPageTemplate('index_next'))
            j=0
    if j!=0:flowables.append(pagebreak)
            


    doc.build(flowables,onFirstPage=index_page, onLaterPages=index_page )

    return outputfile
        
def build_lists(solidbom):
    
    
    listofprocess=[*process_conf]
    
    for process in listofprocess:
        # bom_in=solidbom.flatbom
        # bom_in=bom_in.loc[bom_in['process']==process | bom_in['process2']==process |bom_in['process3']==process]
        process_visual_list(solidbom,process)   
        
        
        

# bomindex(test)






def bom_to_excel(bom_in,outputfolder,title="",qty="qty", firstrow=0):
    
        sheet="Scope"
 
        if title=="": title="List"
        
        excel_file=outputfolder + title +".xlsx"
        
        property_list=['partnumber','revision',qty,'description','material','thickness','finish']
        bom_to_sheet=bom_in[property_list] 


        if firstrow!=0:
            bom_to_sheet.index = np.arange(firstrow, len(bom_to_sheet) + firstrow)
        
        
        #Generate basic lists:
        with pd.ExcelWriter(excel_file) as writer:
            bom_to_sheet.to_excel(writer, sheet_name=sheet)
            
            


 
        
        workbook=load_workbook(filename=excel_file)
        # # #print(dir(self.workbook))
        
        # Add images
        worksheet=workbook[sheet]
        worksheet.insert_cols(idx=2)
        

        
        #Adjust the width of the cols depending on content
        str_flatbom=bom_to_sheet.applymap(str)
        col_width=[]

        #for col in str_flatbom:
        #   col_width.append(str_flatbom[col].map(len).max()*0.8)
        col_width.insert(0,5)
        col_width.insert(1,10)
        col_width.insert(2,25)
        col_width.insert(3,5)
        col_width.insert(4,5)
        col_width.insert(5,40)
        col_width.insert(6,15)
        col_width.insert(7,5)
        col_width.insert(8,15)

        
        
        ##print(col_width)
       
        for i, column_width in enumerate(col_width):
            #if column_width>5:
                worksheet.column_dimensions[get_column_letter(i+1)].width =int( column_width)

 
                

        i=-1 
        
        for index, row in bom_in.iterrows():
            ##print(row)
            i=i+1
            rd =worksheet.row_dimensions[i+2] # get dimension for row 3
            rd.height = 40 # value in points, there is no "auto"
            
            thumb=thumbnail(row.pngpath,
                            size=(100, 100))
            ##print(thumb)
            

            #for j,col in enumerate(maincols):
        
                
            #    if row[col] =='' or str(row[col])=='nan':
            #        cell=chr(67+j)+str(i+2)
            #        if col!="process2" and col!="process3" : 
            #            worksheet[cell].fill=redFill
 
                    
            #Add revision color warning
            if row['revision'] =='':
                worksheet['D'+str(i+2)].fill = yellowFill

            #Add description color warning
            if row['description'] =='':
                worksheet['F'+str(i+2)].fill = yellowFill

            #Add material color warning
            if row['material'] =='':
                worksheet['G'+str(i+2)].fill = yellowFill

            #Add finish color warning
            if row['finish'] =='':
                worksheet['I'+str(i+2)].fill = yellowFill


            # #Add description color warning
            # if self.flatbom.at[i,'description'] =='' or str(self.flatbom.at[i,'description'])=='nan':
            #     worksheet['E'+str(i+2)].fill = redFill
            
            
            
            #Add image
            cell='B'+str(i+2)
            try:
            # if True:
                image=openpyxlIm(thumb)
                image.height=60
                image.width=60
                worksheet.add_image(image, cell)
                
            except:
               # worksheet.add_image(thumb, cell)
               worksheet[cell].fill = redFill
               #print("Could not add image to excel ", row['partnumber'])
               flash("Could not add image to excel ", row['partnumber'])
                
            
        
        #create the border style to put all around the cells
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))   

        #Border annd Center alignment for all cells    
        for row in worksheet.iter_rows():
            for cell in row:
                # #print(dir(cell.style))
                # cell.style.alignment.wrap_text=True
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center',wrapText=True)


        #Put to landscape and adjust the width of the page to the width of the content
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(worksheet, paper_size = int(sum(col_width)), orientation='landscape')

        #Add print title
        worksheet.oddHeader.left.text = title.upper()
        worksheet.oddHeader.left.size = 16

        workbook.save(filename=excel_file)
        
        #Export to pdf
        pdf_path=outputfolder + title +".pdf"
        # excel_to_pdf (excel_file, pdf_path,erasepdf=True)
        
        return pdf_path



# def excel_to_pdf (excel_path, pdf_path,erasepdf=False):

#     o = win32com.client.Dispatch("Excel.Application")
    
#     o.Visible = False
    
#     wb = o.Workbooks.Open(excel_path)
    
#     # ws_index_list = [1] #say you want to print these sheets
    
#     # ws=o.ActiveSheet
#     # wb.WorkSheets(ws).Select()
    
#     wb.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
#     wb.Close(SaveChanges=False)
#     if erasepdf:
#         os.remove(excel_path)


def get_files(bom_in,extension,partfolder,subfolder=True):
     
    if subfolder:
        outputfolder=partfolder+extension+"/"
    else:
        # print(partfolder)
        outputfolder=partfolder+"/"

       
    for index, row in bom_in.iterrows():
        filenamebit=row["partnumber"]+"_REV_"+row["revision"]
        targetfile=""
        

        if extension=='3D-edrawing':
            sourcefile=fileserver_path+deliverables_folder+"edr"+"/"+filenamebit+"."+"eprt"
            targetfile=outputfolder+filenamebit+"_"+row["description"]+"."+"eprt"
            if not os.path.exists(sourcefile):
                sourcefile=fileserver_path+deliverables_folder+"edr"+"/"+filenamebit+"."+"easm"
                targetfile=outputfolder+filenamebit+"_"+row["description"]+"."+"easm"
        else:
            sourcefile=fileserver_path+deliverables_folder+extension+"/"+filenamebit+"."+extension
        
        
        if extension=='dxf':
            targetfile=outputfolder+filenamebit+"-"+row["material"]+"_"+str(row["thickness"])+"mm."+"_"+row["description"]+"."+extension
            # if row["thickness"]==0 or str(row["thickness"])=='':
            #    bom_in.at[index,'notes'].append("Missing thickness")
        elif targetfile=="":
            targetfile=outputfolder+filenamebit+"_"+row["description"]+"."+extension
            #print(targetfile)

        
        try:
            if os.path.exists(sourcefile):
                create_folder_ifnotexists(outputfolder)
            copyfile(sourcefile,targetfile)
        except:
            pass
            #print("source - ", sourcefile)
            #print("target file", targetfile)
            #print("coouldnt get file ", extension," for ",row['partnumber'])


def get_all_files(bom_in,partfolder):

    extensionlist=['pdf','dxf','step','edr']
    feedbackcols=['partnumber','revision','description','process','process2', 'process3']+extensionlist
    
    feedback=bom_in[feedbackcols].copy()

    feedback.to_excel(partfolder+"feedback_"+datetime.now().strftime('%d_%m_%Y-%H_%M_%S_%f')+".xlsx") 



    
    extensionlist=['pdf','dxf','step','3D-edrawing']

    for extension in extensionlist:

        outputfolder=partfolder+extension+"/"
        create_folder_ifnotexists(outputfolder)

        
        for index, row in bom_in.iterrows():
            filenamebit=row["partnumber"]+"_REV_"+row["revision"]
            targetfile=""


            if extension=='3D-edrawing':
                sourcefile=fileserver_path+deliverables_folder+"edr"+"/"+filenamebit+"."+"eprt"
                targetfile=outputfolder+filenamebit+"_"+row["description"]+"."+"eprt"
                if not os.path.exists(sourcefile):
                    sourcefile=fileserver_path+deliverables_folder+"edr"+"/"+filenamebit+"."+"easm"
                    targetfile=outputfolder+filenamebit+"_"+row["description"]+"."+"easm"
            else:
                sourcefile=fileserver_path+deliverables_folder+extension+"/"+filenamebit+"."+extension

            #print("source - ", sourcefile)
            #print("target file", targetfile)

            if extension=='dxf':
                targetfile=outputfolder+filenamebit+"-"+row["material"]+"_"+str(row["thickness"])+"mm."+"_"+row["description"]+"."+extension
                # if row["thickness"]==0 or str(row["thickness"])=='':
                #    bom_in.at[index,'notes'].append("Missing thickness")
            elif targetfile=="":
                targetfile=outputfolder+filenamebit+"_"+row["description"]+"."+extension
                #print(targetfile)
            
            try:
                #print(sourcefile)
                #print(targetfile)
                copyfile(sourcefile,targetfile)
            except:
                pass
                #print("source - ", sourcefile)
                #print("target file", targetfile)
                #print("coouldnt get file ", extension," for ",row['partnumber'])            



def label_page(canvas,doc):
    

    PAGE_HEIGHT = A4[1]
    PAGE_WIDTH = A4[0]
 
    canvas.setFillColor(colors.white)

    canvas.saveState()



def label_list(solidbom, outputfolder="",title=""):
    

    #get children data
    bom_in=solidbom.flatbom




    if outputfolder=="":
        outputfolder=solidbom.folderout


    
    #Remove harddware and sort by process
    #bom_in=bom_in.loc[bom_in['process']!='hardware'].sort_values(by=['process','partnumber'])
    
  
    if title=="":
        outputfile=outputfolder + "Visual_list-"+ solidbom.tag+".pdf"
    else:
        outputfile=outputfolder + title + "-"+ solidbom.tag+".pdf"

    #In case the list is empty
    if len(bom_in)==0:
        return
    
    doc = SimpleDocTemplate(outputfile,    pagesize=A4,
                            rightMargin=5*mm,      leftMargin=5*mm,
                            topMargin=15*mm,      bottomMargin=15*mm)
    pagebreak = PageBreak()
    styles = getSampleStyleSheet()
    
    #Sections building

    
    #Index page template reference
    index_firsframe = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    index_nextframe = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    doc.addPageTemplates([PageTemplate(id='index_first',frames=index_firsframe,pagesize=A4,onPage=label_page),
                         PageTemplate(id='index_next',frames=index_firsframe,pagesize=A4,onPage=label_page),])


    #Create the story for the pdf
    flowables=[]
    
    #Add pdf header
        # headertext=makeParaRight(solidbom.partnumber + " Drawing Pack")
    
    if title=="":
        headertext=makeParaCenter("Visual index" )
    else:
        headertext=makeParaCenter(title)


    # para = Paragraph(headertext, style=styles["Heading1"])
    # flowables.append(para)
    # headertext=makeParaCenter(solidbom.partnumber+  " REV "+solidbom.revision + ": "+solidbom.description)
    # para = Paragraph(headertext, style=styles["Heading2"])
    # flowables.append(para)
    
    
    flowables.append (Spacer(0, 0.1*mm))
    # flowables.append(NextPageTemplate('index_next'))


            
    partlist=Part.partlist(bom_in)
    
    #Split in sections
    boxywidth=90
    boxyheight=38.1
    boxycols=int(((210-20)/ boxywidth))
    margin=15*mm
    spacer=1.5*mm
    boxyrows=int(((297-margin)/ (boxyheight+spacer)))
    

    chunks = [partlist[x:x+boxycols] for x in range(0, len(partlist), boxycols)]
    #flash(boxycols)
   
    #Add all the parts
    j=0
    firstrow=False
    flowables.append (Spacer(0, boxyheight/2*mm+1.5*mm))
    for split_list in chunks:
        if firstrow:
            flowables.append(pagebreak)
            flowables.append (Spacer(0, spacer))
            firstrow=False
            
        grid=BoxyGrid(split_list, width=boxywidth, height=boxyheight,y=-j*boxyheight/2)
        j=j+1
        flowables.append(grid)
        flowables.append (Spacer(0, spacer))
        if j>boxyrows+1:
            # flowables.append(NextPageTemplate('index_next'))
            flowables.append (Spacer(0, boxyheight/2*mm+1.5*mm))
            j=0
            firstrow=True
        


    doc.build(flowables,onFirstPage=index_page, onLaterPages=index_page )

    return outputfile





def loadexcelcompilelist(filein,export_objects=False):
 
    excelfile=pd.ExcelFile(filein)
    partlist=excelfile.parse('COMPILE')
    partlist=partlist.set_index('item')

    partlist['revision']=partlist['revision'].astype(str)
    partlist['revision']=partlist['revision'].str.replace(".0","")
    partlist.loc[partlist.revision=="nan",'revision']=''

    # #print(partlist)
    object_list=[]
    templist=[]
    flatbom=[]
    
    # db.session.refresh()
    #Loop and find parts
    for item, row in partlist.iterrows():
        # #print(row['partnumber'])
        # #print(row['revision'])


        #Find if the part is already in database
        database_part=db.session.query(Part).filter(and_(Part.partnumber==row['partnumber'] ,
                                                    Part.revision==row['revision'])
                                                    ).first()
        if database_part!=None:
            database_part.updatefilespath(webfileserver)
            if 'qty' in partlist.columns:
                # templist.append({'part':database_part,'qty':row['qty']})
                
                if database_part in templist:
                    database_part.qty=row['qty']
                    database_part.totalqty=database_part.totalqty+row['qty']
                else:
                    database_part.totalqty=row['qty']
                    database_part.qty=row['qty']
                    templist.append(database_part)
            else:                
                # templist.append({'part':database_part,'qty':0})
                database_part.qty=0
                templist.append(database_part)
            
        else:
            flash("Couldn't find " + row['partnumber']+" REV " +row['revision'])
        # #print(database_part)
    # for x in templist:
    #     print(x.partnumber, x.qty, x.totalqty)


    # print("temp list",templist)
    
    # #Account for duplicates
    # for item in templist:
        
    #     if len(object_list)>0:
    #         # print(item['part'].partnumber)
    #         # print(len(object_list))
    #         addtolist=False
    #         for finalitem in object_list:
    #             print("$$$", item.partnumber, finalitem.partnumber)
    #             print("$$$", item.qty, finalitem.qty)
    #             print("$$$", item.totalqty, finalitem.totalqty)
                
    #             if item.partnumber==finalitem.partnumber and item.revision==finalitem.revision:
    #                 # item['qty']=item_total['qty']+item['qty']
    #                 # item['part'].qty=item_total['qty']
    #                 print("******finalitem.totalqty******** ",finalitem.totalqty,finalitem.partnumber)
    #                 print("******item.qty******** ",item.qty, item.partnumber)
                    
    #                 finalitem.totalqty=finalitem.totalqty+item.qty+5000000
                    
    #                 print("@@@@@@@@@@@@@ ",finalitem.totalqty)
    #                 db.session.commit()
                    

    #             else: 
    #                 addtolist=True
                 
                   
    #         if addtolist:
    #             item.totalqty=item.qty                
    #             object_list.append(item)
    #             db.session.commit()
    #     else:
    #         item.totalqty=item.qty 
    #         object_list.append(item)
    #         db.session.commit()
    #     db.session.commit()

    # print("object list",object_list)

    # for x in object_list:
    #     partdict=x['part'].as_dict(folder=fileserver_path)
    #     partdict['qty']=x['qty']
    #     flatbom.append(partdict)



    # print("temp list",templist)

    object_list=templist
    for x in object_list:
        x.qty=x.totalqty

    flatbom=pd.DataFrame([x.as_dict(folder=fileserver_path) for x in object_list])



    # object_list=[x['part'] for x in object_list]

    # for x in object_list:
    #     print(x.partnumber, x.qty, x.totalqty)
    # # input("tesaesta")
           
    
    
    # print(flatbom)
    if export_objects:
        return (flatbom,object_list)
    else:
        return(flatbom)




        
class processlengend(Flowable):
    """
    Draw a box + line + text
    -----------------------------------------
    | foobar |
    ---------
    """
    def __init__(self, reflist):
        Flowable.__init__(self)
        self.reflist=reflist
        self.x=10*mm
        self.y=-60*mm*0

        
    def draw(self):

        reflist=self.reflist
        x0=self.x
        y0=self.y

        self.canv.setFont("Helvetica-Bold", 12)
        self.canv.setFillColor(colors.black)
        self.canv.drawString(x0, y0, "Process colour code")

        x=x0
        y=y0-10*mm
        
        self.canv.setFont("Helvetica-Bold", 10)
        self.canv.setLineWidth(3)          

        for processdict in reflist:
            process=processdict['process']
            processdict['rectwidth']=(len(processdict['process'])*2+3)*mm
            

            #Split lines
            if  x+processdict['rectwidth'] >500:
                x=x0
                y=y-10*mm
            
                
            #Add text with colour
            self.canv.setFillColor(get_process_color(process))
            self.canv.setStrokeColor(get_process_color(process))
            self.canv.drawString(x, y, process)
            self.canv.rect(-1.5*mm+x, -1.5*mm+y, processdict['rectwidth'], 6*mm,fill=0)     

            x=x+processdict['rectwidth']+2*mm 
            
            
            
def addPageNumbersOnly (filein,fileout,index=False):

    if os.path.isfile(filein):

        pdfWriter = PyPDF2.PdfFileWriter()
        pdfFileObj = open(filein,'rb')
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

        totalpages=pdfReader.numPages

        #Opening each page of the PDF
        for pageNum in range(pdfReader.numPages):
            pageObj = pdfReader.getPage(pageNum)
                            
            pageObj=pdf_pagenum(pageObj,pageNum-1*index, total=totalpages)
        
            pdfWriter.addPage(pageObj)
            

        with open(fileout, 'wb') as fh:
            pdfWriter.write(fh)
        
        

def partTable(reflist):

    

    data=[]
    header=['partnumber', 'description','qty']
    data.append([x.upper() for x in header])

    for part in reflist:
        for field in header:
            rowdata=[part.__dict__[field] for field in header]
        data.append(rowdata)

    t=Table(data)
    t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),
                            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                            ('FONTSIZE',(0,0),(-1,-1),3*mm),
                            ]))

    return t

    
        







            

